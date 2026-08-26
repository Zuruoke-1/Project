"""
prism_xml_method.py — Write data to Prism template via XML, open Prism, extract results.

This approach:
1. Reads .txt data files directly (avoids Excel formula issues)
2. Writes X,Y data into the .pzfx template via XML manipulation
3. Opens Prism to run the pre-configured analysis
4. Extracts k0, k1, tau, viscosity results back to Excel
5. Trims Data Input to Matlab to shortest unempty range (for easier Matlab processing)
6. Writes trimmed Data input 1 and input 2 to file 2 and file 1 respectively (for Matlab to read)
7. Runs moduliloop.4_edited which results in Allbeads file which is copied to excel.
8. Optinonally, can run in batch mode to process multiple conditions  or set batch mode to False to process a single conditions file.

Install:
    pip install pandas openpyxl pyperclip xlwings more-itertools scipy

Usage:
    python prism_xml_method.py
"""


from pathlib import Path
import pandas as pd
import os
import re
import shutil
import subprocess
import time
import tempfile
import xlwings as xw
from openpyxl import load_workbook
import math
# from more_itertools import islice, batched
# import itertools 
from itertools import islice, batched
import sys
sys.path.insert(0, "/Users/sharonnyoyoko")
# sys.path.insert(0, "/Users/fntf0423")

import xml.etree.ElementTree as ET
from pipeline import *
import colprint as cp


# ═════════════════════════════════════════════════════════════=
#  CONFIGURATION — edit these paths
# ═════════════════════════════════════════════════════════════=
# Use raw strings for Windows paths to avoid unicode-escape errors


'''
To make file path handling easier, I have put all the necessary files in a folder called automation requirements.
So just adjust the path to the Automation requirements folder and various software and the rest of the code should work.
'''

# ══════════════════════════════════════════════════════════════
#  CONFIG — EDIT THESE PATHS
# ══════════════════════════════════════════════════════════════
# Every changeable path is listed here. Edit the values in this script.

# Folder containing the Excel template, Prism bead template, etc.
AUTOMATION_REQUIREMENTS_PATH = "/Users/sharonnyoyoko/Downloads/redone anlysis/Automation Requirements"
# Example (Windows):  AUTOMATION_REQUIREMENTS_PATH = r"C:\Users\fntf0423\Downloads\Analysis\Automation Requirements"

# The Excel template used to build each sample's workbook.
EXCEL_TEMPLATE = os.path.join(AUTOMATION_REQUIREMENTS_PATH, "Clotting and lysis TEMPLATE_25Hz_1min pulls.xlsx")

# The Prism bead template that receives the raw X/Y data.
PZFX_TEMPLATE = os.path.join(AUTOMATION_REQUIREMENTS_PATH, "MagTweezers Analysis GRAPHPAD TEMPLATE.pzfx")

ANALYSIS_DIR = AUTOMATION_REQUIREMENTS_PATH
PRISM_APP = "Prism 10"   # "Prism 11" on some Mac setups

# Single-sample folder — only used when BATCH_MODE is off (batch mode overrides it).
FOLDER_PATH = "/Users/sharonnyoyoko/Downloads/redone anlysis/20251024_ami21/AMI21/p1b2 separated"

# ---- Batch processing (edit per experiment) ----
BATCH_MODE = True          # True = process all samples and build grouped G' tables
BASE_DIR_FOR_BATCH = "/Users/sharonnyoyoko/Downloads/g/20260313"

GP_SUMMARY_FILENAME = "Gp_summary.csv"  # single grouped grid file: rows=pulls, cols=samples, G'/G'' blocks

# grid block labels: wide-CSV column prefix -> block header in Gp_summary.csv
GP_GRID_LABELS = {
    "gp01_": "G' 0.1Hz", "gp1_": "G' 1Hz", "gp10_": "G' 10Hz",
    "gdp01_": "G'' 0.1Hz", "gdp1_": "G'' 1Hz", "gdp10_": "G'' 10Hz",
    "td01_": "tanδ 0.1Hz", "td1_": "tanδ 1Hz", "td10_": "tanδ 10Hz",
}


# ═════════════════════════════════════════════════════════════=
#  FIXED CONFIGURATION — do not edit below this line (except necessary)
# ═════════════════════════════════════════════════════════════=

'''Confirm that the following paths exist and are correct for your system.'''

file2_dir = os.path.join(ANALYSIS_DIR, "file2.txt")
file1_dir = os.path.join(ANALYSIS_DIR, "file1.txt")

INPUT_SHEET = 'Data Input from Text File'
RESULTS_CELL = 'Q2'  # Where to write k0, k1, tau, viscosity in Excel
WANTED_PARAMS = ["k0", "k1", "tau", "viscosity"]
BEAD_COUNT = 10
# PRISM_APP = "Prism 10" #10 for windows /11 for mac
PRISM_LOAD_DELAY = 15  # seconds to wait for Prism to open
PRISM_HEADLESS = False   # run Prism hidden in background (no window focus)
PRISM_POLL_TIMEOUT = 180   # max seconds to poll for analysis completion
PRISM_POLL_INTERVAL = 5    # seconds between save-and-check cycles
# ── Platform-specific config ──
if sys.platform == "darwin":
    PRISM_APP_PATH = "/Applications/Prism 10.app"
    MATLAB_CMD = "/Applications/MATLAB_R2025b.app/bin/matlab"
    PRISM_SCRIPT_DIR = os.path.join(AUTOMATION_REQUIREMENTS_PATH, "Bead results")
    done_file = os.path.join(PRISM_SCRIPT_DIR, "done.txt")
elif sys.platform == "win32":
    PRISM_APP_PATH = r"C:\Program Files\GraphPad\Prism 10\prism.exe"
    possible_matlab_paths = [
        r"C:\Program Files\MATLAB\R2025a\bin\matlab.exe",
        r"C:\Program Files\MATLAB\R2024a\bin\matlab.exe",
        r"C:\Program Files\MATLAB\R2023a\bin\matlab.exe",
    ]
    MATLAB_CMD = next((p for p in possible_matlab_paths if os.path.exists(p)), "matlab")
    if MATLAB_CMD == "matlab":
        cp.warn("WARNING: No MATLAB executable found in standard Windows paths; will try 'matlab' from PATH")
    else:
        print(f"Using MATLAB executable: {MATLAB_CMD}")
    PRISM_SCRIPT_DIR = os.path.join(AUTOMATION_REQUIREMENTS_PATH, "Bead results")
    done_file = os.path.join(PRISM_SCRIPT_DIR, "done.txt") 
    
else:
    PRISM_APP_PATH = None
    MATLAB_CMD = "matlab"
    PRISM_SCRIPT_DIR = os.path.expanduser("~/Downloads/Bead results")
    done_file = os.path.join(PRISM_SCRIPT_DIR, "done.txt")


GP_SHEET_NAME = 'For GraphPad Graph&Stats'
GP_POINTONE_RANGE = 'B15:B24'
GP_ONE_RANGE = 'C15:C24'
GP_TEN_RANGE = 'D15:D24'

GDP_POINTONE_RANGE = 'E15:E24'
GDP_ONE_RANGE = 'F15:F24'
GDP_TEN_RANGE = 'G15:G24'
# tanδ = G''/G' — computed by the Excel template (source of truth). The
# pipeline falls back to computing G''/G' when these cells are missing/zero.
TAND_POINTONE_RANGE = 'H15:H24'
TAND_ONE_RANGE = 'I15:I24'
TAND_TEN_RANGE = 'J15:J24'
# TIME_RANGE = 'A15:A24'


# ══════════════════════════════════════════════════════════════
#  PRISM NAMESPACE
# ══════════════════════════════════════════════════════════════

_NS = "http://graphpad.com/prism/Prism.htm"
ET.register_namespace("", _NS)
ET.register_namespace("dt", "urn:schemas-microsoft-com:datatypes")


def ns(tag):
    """Return namespace-qualified tag name."""
    return f"{{{_NS}}}{tag}"


# ══════════════════════════════════════════════════════════════
#  HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════

def natural_sort_key(f):
    parts = re.split(r'(\d+)', f.name)
    return [int(p) if p.isdigit() else p.lower() for p in parts]


def is_windows():
    return sys.platform.startswith("win")


def is_macos():
    return sys.platform == "darwin"


def validate_xy_columns(df):
    """Ensure the DataFrame has X and Y1 in the first two columns and no
    substantive data in later columns.

    Returns (ok, reason, trimmed_df) where `ok` is True when validation
    passes, `reason` is a message, and `trimmed_df` is a DataFrame containing
    only the first two columns (for safe writing).
    """
    if df.shape[1] < 2:
        return False, "Less than two columns (need X and Y1)", df

    # Consider a value substantive if it's not NA and not an empty/whitespace string
    def has_substantive(s):
        if pd.isna(s):
            return False
        if isinstance(s, str) and s.strip() == "":
            return False
        return True

    extras = df.iloc[:, 2:]
    if not extras.empty:
        # Check any substantive value in extra columns (per-column check)
        for col in extras.columns:
            if extras[col].apply(has_substantive).any():
                return False, "Data found in columns beyond the first two (only X and Y1 allowed)", df.iloc[:, :2].copy()

    # Also ensure first two columns contain at least one substantive value
    first_two = df.iloc[:, :2]
    found12 = False
    for col in first_two.columns:
        if first_two[col].apply(has_substantive).any():
            found12 = True
            break
    if not found12:
        return False, "No data found in first two columns (X and Y1)", first_two

    return True, "OK", first_two.copy()


def get_number(filename):
    match = re.search(r'(\d+)', filename)
    return int(match.group(1)) if match else None


def normalize_used_range(values):
    if values is None:
        return []
    if not isinstance(values, list):
        values = [[values]]
    if not isinstance(values[0], list):
        values = [values]

    rows = [list(row) for row in values]

    def is_empty_row(row):
        return all(
            v is None or (isinstance(v, str) and v.strip() == "")
            for v in row
        )

    while rows and is_empty_row(rows[0]):
        rows.pop(0)
    while rows and is_empty_row(rows[-1]):
        rows.pop()

    if not rows:
        return []

    max_col = max(len(row) for row in rows)
    def is_empty_cell(v):
        return v is None or (isinstance(v, str) and v.strip() == "")

    non_empty_columns = [
        idx for idx in range(max_col)
        if any(not is_empty_cell(row[idx]) for row in rows if idx < len(row))
    ]

    if not non_empty_columns:
        return []

    return [
        [row[idx] if idx < len(row) else None for idx in non_empty_columns]
        for row in rows
    ]


def build_excel_output_path(folder_path, bead_range):
    """Return an Excel output path using the condition folder name."""

    sample_name = folder_path.name.replace(" separated", "").replace(".separated", "")
    condition = folder_path.parent.name
    if condition and not sample_name.startswith(f"{condition}_"):
        output_name = f"{condition}_{sample_name}_bead{bead_range}.xlsx"
    else:
        output_name = f"{sample_name}_bead{bead_range}.xlsx"
    return folder_path.parent / output_name


def read_range(ws, cell_range):
    """Read a worksheet range and return a flat list of values."""
    return [cell.value for row in ws[cell_range] for cell in row]


def read_gp_values_from_excel(excel_path):
    """Read G' values from the 'For GraphPad Graph&Stats' sheet."""
    wb = load_workbook(excel_path, data_only=True)
    if GP_SHEET_NAME not in wb.sheetnames:
        wb.close()
        return None, None, None, None, None, None, None, None, None

    ws = wb[GP_SHEET_NAME]
    gp01 = read_range(ws, GP_POINTONE_RANGE)
    gp1 = read_range(ws, GP_ONE_RANGE)
    gp10 = read_range(ws, GP_TEN_RANGE)
    gdp01 = read_range(ws, GDP_POINTONE_RANGE)
    gdp1 = read_range(ws, GDP_ONE_RANGE)
    gdp10 = read_range(ws, GDP_TEN_RANGE)
    td01 = read_range(ws, TAND_POINTONE_RANGE)
    td1 = read_range(ws, TAND_ONE_RANGE)
    td10 = read_range(ws, TAND_TEN_RANGE)
    wb.close()
    print(f""" Gp_pointone = {gp01}, Gp_one={gp1}, Gp_ten={gp10}
            Gdoublep_pointone = {gp01}, Gdoublep_one={gp1}, Gdoublep_ten={gp10}""")
    return gp01, gp1, gp10, gdp01, gdp1, gdp10, td01, td1, td10 


# (old append_gp_summary removed — superseded by append_gp_and_gdp_summary below)


def read_gp_summary_grid(grid_path):
    """Read the grouped Gp_summary.csv back as {block_label: DataFrame}.

    Each block DataFrame: rows = pull numbers (index 'pull'),
    columns = 'condition sample'.
    """
    if not os.path.exists(grid_path):
        return {}
    df = pd.read_csv(grid_path, encoding="utf-8-sig", header=[0, 1], index_col=0)
    df.index.name = "pull"
    blocks = {}
    for label in df.columns.get_level_values(0).unique():
        blocks[str(label)] = df[label]
    return blocks


def normalize_gp_summary(wide):
    """Merge duplicate rows (same pull under different labels) and re-label
    every row to the file's current scheme: position = pull - min_pull + 1.

    Lossless: values from the non-preferred variant fill gaps. Also drops
    all-empty rows and sorts by pull. Every write normalizes through here,
    so re-runs / mixed sample starts can never leave the file messy.
    """
    pulls = []
    for lbl in wide.index:
        try:
            pulls.append(int(str(lbl).split("_")[-1]))
        except Exception:
            pulls.append(None)
    valid = [p for p in pulls if p is not None]
    if not valid:
        return wide
    min_pull = min(valid)

    variants = {}
    for lbl, row in wide.iterrows():
        try:
            p = int(str(lbl).split("_")[-1])
        except Exception:
            continue
        variants.setdefault(p, {})[str(lbl)] = row

    merged = {}
    for p, rows in variants.items():
        current = f"{p - min_pull + 1}_{p}"
        base_lbl = current if current in rows else list(rows)[0]
        base = rows[base_lbl].copy()
        for lbl, row in rows.items():
            if lbl == base_lbl:
                continue
            base = base.fillna(row)          # fill gaps from the other variant
        merged[p] = (current, base)

    order = sorted(merged)
    out = pd.concat([merged[p][1].rename(merged[p][0]) for p in order], axis=1).T
    out.index.name = "pull"
    return out.dropna(how="all")


def upsert_sample_in_gp_summary(grid_path, condition, sample_name, first_num,
                                values_by_prefix):
    """Add or update ONE sample in the single grouped Gp_summary.csv.

    values_by_prefix: {"gp01_": [v0..vN-1], "gdp10_": [...], ...} where the
    i-th value belongs to pull (first_num + i). Read-modify-write: the whole
    grid is reloaded, the sample's old columns are dropped, the new ones are
    added, and the file is rewritten.

    Row labels are "{position}_{pull}" — position counts beads continuously
    across batches (1..10, then 11..20, ...), anchored to the smallest pull
    in the file: position = pull - min_pull + 1. A file starting at pull 2
    gives "1_2".."10_11", "11_12".."20_21"; bead3-12 still starts "1_3".
    """
    sample_col = f"{condition} {sample_name}"
    lengths = [len(v) for v in values_by_prefix.values() if v]
    if not lengths:
        cp.warn(f"  WARNING: no G'/G'' values for {condition}/{sample_name} — nothing to write")
        return
    n = max(lengths)

    # Guard: a valid pull batch must start at pull >= 1. first_num = 0 (e.g. a
    # file named "pulling0.txt" / "0.txt", or a name whose first digits are 0)
    # would produce garbage labels like "-1_0", "0_1" — refuse instead.
    if first_num < 1:
        cp.warn(f"  WARNING: first pull number {first_num} is invalid (< 1) for "
                f"{condition}/{sample_name} — skipping summary write. "
                f"Check the pulling file names (they must start at pull >= 1).")
        return

    blocks = read_gp_summary_grid(grid_path)

    # Anchor: the smallest pull number anywhere in the grid (or this batch).
    # Position = pull - min_pull + 1, so batches of every sample number
    # continuously: first batch 1..10, second batch 11..20, ...
    min_pull = first_num
    for block in blocks.values():
        for lbl in block.index:
            try:
                p = int(str(lbl).split("_")[-1])
                if p > 0:
                    min_pull = min(min_pull, p)
            except Exception:
                continue

    labels = [f"{first_num + i - min_pull + 1}_{first_num + i}" for i in range(n)]  # informational
    for prefix, label in GP_GRID_LABELS.items():
        vals = values_by_prefix.get(prefix)
        if not vals:
            continue
        block = blocks.get(label)
        if block is None:
            block = pd.DataFrame(index=pd.Index([], name="pull"))
        # Merge by PULL (the number after "_"), never by the full label — so a
        # label-scheme change between runs can't duplicate the same pull.
        by_pull = {first_num + i: vals[i] for i in range(len(vals))}
        if sample_col in block.columns:
            old_series = block[sample_col].dropna()
            for lbl, val in old_series.items():
                try:
                    p = int(str(lbl).split("_")[-1])
                except Exception:
                    continue
                by_pull.setdefault(p, val)   # keep pulls not in this batch
            block = block.drop(columns=[sample_col])
        # Re-label EVERYTHING to the current scheme: position = pull - min_pull + 1
        relabelled = {f"{p - min_pull + 1}_{p}": v for p, v in by_pull.items() if p >= min_pull}
        new_col = pd.Series(relabelled, name=sample_col)
        block = pd.concat([block, new_col.to_frame()], axis=1)
        block = block[sorted(block.columns, key=str)]  # keep sample columns tidy
        blocks[label] = block

    if not blocks:
        cp.warn("  WARNING: no G'/G'' values to write")
        return

    wide = pd.concat(list(blocks.values()), axis=1, keys=list(blocks.keys()))
    wide = normalize_gp_summary(wide)
    wide.to_csv(grid_path, encoding="utf-8-sig", index_label="pull")
    print(f"  Updated {grid_path} with {condition}/{sample_name}")

    # keep only ONE summary file per condition folder
    legacy = os.path.join(os.path.dirname(grid_path), "Gp_values.csv")
    if os.path.exists(legacy):
        try:
            os.remove(legacy)
            cp.info(f"  Removed legacy {legacy} (replaced by {GP_SUMMARY_FILENAME})")
        except OSError:
            pass


# Redoing append_gp_summary
def append_gp_and_gdp_summary(folder_path, excel_path, first_num, last_num, bead_count=BEAD_COUNT):
    """Append a sample's G' AND G'' values to the per-condition summary.

    Writes two files into the condition folder:
      Gp_values.csv        wide machine format (one row per sample) — kept for
                           fast appends and total_result.py compatibility.
      Gp_summary_grid.csv  grouped grid (one row per pull, one column per
                           sample, G'/G'' frequency blocks) — opens in Excel
                           looking like the summary sheet.
    """
    condition = folder_path.parent.name
    sample_name = folder_path.name.replace(" separated", "").replace(".separated", "")
    gp01, gp1, gp10, gdp01, gdp1, gdp10, td01, td1, td10 = read_gp_values_from_excel(excel_path)
    # If openpyxl didn't find the sheet or values look empty/zero, try reading
    # live from Excel via xlwings (works when workbook is still open or Excel
    # has recently calculated formulas).
    def looks_all_zero(arr):
        if arr is None:
            return True
        try:
            return all((v == 0 or v == 0.0 or v is None or str(v).strip() in ("", "0", "0.0")) for v in arr)
        except Exception:
            return False

    if gp01 is None or (looks_all_zero(gp01) and looks_all_zero(gp1) and looks_all_zero(gp10)):
        try:
            import xlwings as _xw
            app = _xw.App(visible=False)
            wb = app.books.open(str(excel_path))
            try:
                # Force full recalculation before reading formula results.
                try:
                    wb.api.Application.CalculateFullRebuild()
                except Exception:
                    try:
                        app.api.CalculateFullRebuild()
                    except Exception:
                        pass

                ws = wb.sheets[GP_SHEET_NAME]
                gp01_live = ws.range(GP_POINTONE_RANGE).options(ndim=1).value
                gp1_live = ws.range(GP_ONE_RANGE).options(ndim=1).value
                gp10_live = ws.range(GP_TEN_RANGE).options(ndim=1).value

                gdp01_live = ws.range(GDP_POINTONE_RANGE).options(ndim=1).value
                gdp1_live = ws.range(GDP_ONE_RANGE).options(ndim=1).value
                gdp10_live = ws.range(GDP_TEN_RANGE).options(ndim=1).value

                td01_live = ws.range(TAND_POINTONE_RANGE).options(ndim=1).value
                td1_live = ws.range(TAND_ONE_RANGE).options(ndim=1).value
                td10_live = ws.range(TAND_TEN_RANGE).options(ndim=1).value
                
                # normalize to lists of length 10
                gp01 = gp01_live if isinstance(gp01_live, (list, tuple)) else [gp01_live]
                gp1 = gp1_live if isinstance(gp1_live, (list, tuple)) else [gp1_live]
                gp10 = gp10_live if isinstance(gp10_live, (list, tuple)) else [gp10_live]

                gdp01 = gdp01_live if isinstance(gdp01_live, (list, tuple)) else [gdp01_live]
                gdp1 = gdp1_live if isinstance(gdp1_live, (list, tuple)) else [gdp1_live]
                gdp10 = gdp10_live if isinstance(gdp10_live, (list, tuple)) else [gdp10_live]

                td01 = td01_live if isinstance(td01_live, (list, tuple)) else ([td01_live] if td01_live is not None else [])
                td1 = td1_live if isinstance(td1_live, (list, tuple)) else ([td1_live] if td1_live is not None else [])
                td10 = td10_live if isinstance(td10_live, (list, tuple)) else ([td10_live] if td10_live is not None else [])
            finally:
                try:
                    wb.save()
                    wb.close()
                except Exception:
                    pass
                app.quit()
            print(f"  Read live G' values from Excel for {sample_name}")
        except Exception as exc:
            print(f"  Could not read live Excel values for {sample_name}: {exc}")

    if gp01 is None:
        cp.warn(f"  WARNING: '{GP_SHEET_NAME}' not found in {excel_path}; skipping G' summary.")
        return

    if (gp01 is None or gp1 is None or gp10 is None
            or gdp01 is None or gdp1 is None or gdp10 is None
            or len(gp01) != 10 or len(gp1) != 10 or len(gp10) != 10
            or len(gdp01) != 10 or len(gdp1) != 10 or len(gdp10) != 10):
        cp.warn(f"  WARNING: Unexpected G' range length for {sample_name}; skipping summary. gp01={gp01} gp1={gp1} gp10={gp10}... gdp01={gdp01} gdp1={gdp1} gdp10={gdp10}")
        return

    # refuse to write all-zero/empty values (means the values were never
    # computed/saved in the workbook — writing zeros would fake real data)
    if (looks_all_zero(gp01) and looks_all_zero(gp1) and looks_all_zero(gp10)
            and looks_all_zero(gdp01) and looks_all_zero(gdp1) and looks_all_zero(gdp10)):
        cp.warn(f"  WARNING: no usable G'/G'' values for {sample_name} "
              f"(all zero/empty in {excel_path}) — skipping summary write")
        return

    # Trailing beads missing (e.g. bead 10)? Write their G' as nil (blank)
    # instead of whatever 0/empty value the workbook holds for them.
    if bead_count < 10:
        def _nil_tail(lst):
            if lst is None:
                return lst
            lst = list(lst)                      # xlwings may return tuples
            for i in range(bead_count, len(lst)):
                lst[i] = None
            return lst
        gp01, gp1, gp10 = _nil_tail(gp01), _nil_tail(gp1), _nil_tail(gp10)
        gdp01, gdp1, gdp10 = _nil_tail(gdp01), _nil_tail(gdp1), _nil_tail(gdp10)
        td01, td1, td10 = _nil_tail(td01), _nil_tail(td1), _nil_tail(td10)

    # tanδ read from the workbook may come back short/missing (old files or
    # sheets without the tanδ columns); pad to length 10 to stay aligned.
    def _pad10(lst):
        lst = list(lst) if lst else []
        return lst + [None] * (10 - len(lst)) if len(lst) < 10 else lst
    td01, td1, td10 = _pad10(td01), _pad10(td1), _pad10(td10)

    # Excel's tanδ (H15:J24) is the source of truth; when a cell is missing
    # or zero there, fall back to computing G''/G' (error handling).
    def _resolve_td(gp_l, gdp_l, td_l):
        out = []
        for g, gd, t in zip(gp_l, gdp_l, td_l):
            if t is None or t == 0:
                try:
                    g, gd = float(g), float(gd)
                    t = gd / g if g and g != 0 else None
                except Exception:
                    t = None
            out.append(t)
        return out
    td01 = _resolve_td(gp01, gdp01, td01)
    td1 = _resolve_td(gp1, gdp1, td1)
    td10 = _resolve_td(gp10, gdp10, td10)

    add_number = first_num - 1
    n = min(last_num, len(gp01))   # never index past the 10 values

    # write/update the single grouped Gp_summary.csv (rows=pulls, cols=samples)
    values_by_prefix = {
        "gp01_": [gp01[i] for i in range(n)],
        "gp1_": [gp1[i] for i in range(n)],
        "gp10_": [gp10[i] for i in range(n)],
        "gdp01_": [gdp01[i] for i in range(n)],
        "gdp1_": [gdp1[i] for i in range(n)],
        "gdp10_": [gdp10[i] for i in range(n)],
        "td01_": [td01[i] for i in range(n)],
        "td1_": [td1[i] for i in range(n)],
        "td10_": [td10[i] for i in range(n)],
    }
    upsert_sample_in_gp_summary(
        folder_path.parent / GP_SUMMARY_FILENAME,
        condition, sample_name, first_num, values_by_prefix)


def trim_matlab_columns(data, bead_count):
    """Keep only the first bead_count bead pairs (2 columns each) of a range.

    Used when trailing beads are missing (e.g. bead 10): MATLAB then sees
    exactly the beads that exist instead of empty/zero columns.
    """
    ncols = 2 * bead_count
    if isinstance(data, list) and data:
        if isinstance(data[0], list):
            return [row[:ncols] for row in data]
        return data[:ncols]
    return data


def write_matlab_input(path, values):
    rows = normalize_used_range(values)
    bad = []
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            if value is None or (isinstance(value, str) and value.strip() == ""):
                continue
            if isinstance(value, (int, float)):
                continue
            if isinstance(value, str):
                if re.fullmatch(r"[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?", value.strip()):
                    continue
            bad.append((r_idx, c_idx, value))
    if bad:
        cp.warn(f"  WARNING: {path} contains non-numeric values at {len(bad)} positions. "
              f"First entries: {bad[:10]}")

    row_lengths = [len(row) for row in rows]
    if len(set(row_lengths)) > 1:
        cp.warn(f"  WARNING: {path} has uneven row lengths: {sorted(set(row_lengths))}")

    with open(path, "w", encoding="utf-8") as f:
        for row in rows:
            f.write("\t".join("" if v is None else str(v) for v in row))
            f.write("\n")

    print(f"  Wrote {len(rows)} rows x {row_lengths[0] if rows else 0} cols to {path}")
    return rows



def load_txt_data(folder: Path):
    files = sorted(folder.glob("*.txt"))
    files = [f for f in files if "pulling" in f.name]

    dataframes = []
    for f in files[:BEAD_COUNT]:
        df = pd.read_csv(f, sep="\t")
        dataframes.append(df)
    return dataframes

def write_excel_template(dataframes, template_path, output_path):
    app = xw.App(visible=False)
    wb = app.books.open(str(template_path))
    ws = wb.sheets["Data Input from Text File"]

    col = 1
    for df in dataframes:
        ws.cells(2, col).value = df.values.tolist()
        col += 6

    wb.save(str(output_path))
    wb.close()
    app.quit()

    return output_path

def write_data_to_prism(root, table_name, df):
    """Write a DataFrame into an XY table in the .pzfx tree.
    
    Column 0 of df → XColumn, Column 1+ → YColumn subcolumns.
    Operates on the existing root element in-memory.
    """
    target_table = None
    for table in root.iter(ns("Table")):
        title_el = table.find(ns("Title"))
        if title_el is not None and title_el.text == table_name:
            target_table = table
            break

    if target_table is None:
        cp.err(f"  ERROR: Table '{table_name}' not found")
        return False

    table_type = target_table.get("TableType", "")

    if table_type == "XY":
        x_col = target_table.find(ns("XColumn"))
        y_col = target_table.find(ns("YColumn"))

        # Create YColumn if missing
        if y_col is None:
            y_col = ET.SubElement(target_table, ns("YColumn"))
            y_col.set("Width", "81")
            y_col.set("Subcolumns", "1")
            y_col.set("Decimals", "4")
            ET.SubElement(y_col, ns("Title"))
            print("Ycolumn has been created")

        print(ET.tostring(x_col, encoding="unicode"))

        # Write column 0 → XColumn AND XAdvancedColumn Subcolumns
        x_subcol = x_col.find(ns("Subcolumn")) if x_col is not None else None 
        x_adv = target_table.find(ns("XAdvancedColumn"))
        x_adv_subcol = x_adv.find(ns("Subcolumn")) if x_adv is not None else None

        if x_subcol is not None:
            old_d = x_subcol.findall(ns("d"))
            for d in old_d:
                x_subcol.remove(d)
            x_values = df.iloc[:, 0].dropna().tolist()
            for value in x_values:
                d = ET.SubElement(x_subcol, ns("d"))
                d.text = str(value)
            print(f"  XColumn: wrote {len(x_values)} values")

        # Also write to XAdvancedColumn
        if x_adv is not None:
            x_adv_subcol = x_adv.find(ns("Subcolumn"))
            if x_adv_subcol is not None:
                old_d = x_adv_subcol.findall(ns("d"))
                for d in old_d:
                    x_adv_subcol.remove(d)
                x_values = df.iloc[:, 0].dropna().tolist()
                for value in x_values:
                    d = ET.SubElement(x_adv_subcol, ns("d"))
                    d.text = str(value)
                print(f"  XAdvancedColumn: wrote {len(x_values)} values")

        
       
        # Write remaining columns → YColumn Subcolumns
        y_subcols = y_col.findall(ns("Subcolumn")) if y_col is not None else []
        for i in range(1, len(df.columns)):
            subcol_idx = i - 1
            if subcol_idx < len(y_subcols):
                subcol = y_subcols[subcol_idx]
                for d in subcol.findall(ns("d")):
                    subcol.remove(d)
            else:
                subcol = ET.SubElement(y_col, ns("Subcolumn"))
            y_values = df.iloc[:, i].dropna().tolist()
            for value in y_values:
                d = ET.SubElement(subcol, ns("d"))
                d.text = str(value)
            print(f"  YColumn: wrote {len(y_values)} values (first={y_values[0] if y_values else 'empty'})")
    else:
        # Generic table format
        existing_subcols = target_table.findall(ns("Subcolumn"))
        for i, col_name in enumerate(df.columns):
            if i < len(existing_subcols):
                subcol = existing_subcols[i]
                for d in subcol.findall(ns("d")):
                    subcol.remove(d)
            else:
                subcol = ET.SubElement(target_table, ns("Subcolumn"))
            for value in df[col_name].dropna():
                d = ET.SubElement(subcol, ns("d"))
                d.text = str(value)

    print(f"  Wrote {len(df.columns)} cols ({len(df)} rows) into '{table_name}'")
    return True


def save_pzfx(tree, path):
    """Save .pzfx with format Prism expects (encoding, line endings)."""
    tree.write(path, xml_declaration=True, encoding="utf-8")

    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    # Fix XML declaration to match Prism format
    content = content.replace(
        "<?xml version='1.0' encoding='utf-8'?>",
        '<?xml version="1.0" encoding="UTF-8"?>'
    )
    # Fix namespace prefix
    content = content.replace('xmlns:ns1=', 'xmlns:dt=')
    content = content.replace('ns1:', 'dt:')
    # Add Windows carriage returns
    content = content.replace("\n", "\r\n")

    with open(path, "w", encoding="utf-8") as f:
        f.write(content)


def validate_prism_table_columns(root, table_name):
    """Validate that in the Prism `Table` named `table_name` the XColumn has
    data and only the first Y subcolumn contains data (no substantive data in
    later Y subcolumns).

    Returns (ok, message).
    """
    target_table = None
    for table in root.iter(ns("Table")):
        title_el = table.find(ns("Title"))
        if title_el is not None and title_el.text == table_name:
            target_table = table
            break

    if target_table is None:
        return False, f"Table '{table_name}' not found in .pzfx"

    x_col = target_table.find(ns("XColumn"))
    if x_col is None:
        return False, "Missing XColumn"
    x_sub = x_col.find(ns("Subcolumn"))
    if x_sub is None or len(x_sub.findall(ns("d"))) == 0:
        return False, "XColumn has no data"

    y_col = target_table.find(ns("YColumn"))
    if y_col is None:
        return False, "Missing YColumn"
    y_subcols = y_col.findall(ns("Subcolumn"))
    if len(y_subcols) == 0:
        return False, "YColumn has no subcolumns/data"

    # Check first Y subcolumn has data
    first = y_subcols[0]
    if len(first.findall(ns("d"))) == 0:
        return False, "First Y subcolumn has no data"

    # Ensure later Y subcolumns contain no substantive data
    for extra in y_subcols[1:]:
        for d in extra.findall(ns("d")):
            if d.text and d.text.strip() != "":
                return False, "Data found in Y subcolumns beyond the first"

    return True, "Prism table columns validated"


def wait_for_prism_save(prism_path, timeout=90, poll_interval=2):
    """Poll until the .prism file exists AND is a valid, fully-written zip."""
    import zipfile
    elapsed = 0
    print(f"  Waiting for Prism to save .prism file (up to {timeout}s)...")
    while elapsed < timeout:
        if os.path.exists(prism_path):
            try:
                with zipfile.ZipFile(prism_path) as z:
                    z.testzip()
                print(f"  .prism file ready after {elapsed}s")
                return True
            except zipfile.BadZipFile:
                pass  # exists but still being written — keep waiting
        time.sleep(poll_interval)
        elapsed += poll_interval
    return False


def _prism_pids():
    """Return the set of PIDs for running Prism processes (macOS only)."""
    if sys.platform != "darwin":
        return set()
    try:
        out = subprocess.run(["pgrep", "-f", "Prism"],
                             capture_output=True, text=True, timeout=5).stdout
    except Exception:
        return set()
    pids = set()
    for tok in out.split():
        try:
            pids.add(int(tok))
        except ValueError:
            pass
    return pids


def cleanup_prism_recovery():
    """Remove Prism crash-recovery leftovers so the next launch starts clean.

    Prism keeps a per-run backup directory at:
        ~/Library/Application Support/GraphPad/Prism/<version>/run/<pid>/
    plus lock files <version>/.pzr*.tmp
    When Prism is force-killed these survive, and the next launch reopens them
    as 'not well closed' documents (the '(Autosaved)' files that pile up in the
    sample folders). Delete any run dir whose PID is no longer alive and clear
    stale lock files — but NEVER touch a live instance's data.
    """
    if sys.platform != "darwin":
        return
    import glob as _glob
    app_support = os.path.expanduser("~/Library/Application Support/GraphPad/Prism")
    if not os.path.isdir(app_support):
        return
    live = _prism_pids()
    removed_locks = 0
    removed_dirs = 0
    for ver in sorted(os.listdir(app_support)):
        ver_dir = os.path.join(app_support, ver)
        if not os.path.isdir(ver_dir):
            continue
        # Lock files — only safe to remove when no Prism instance is running
        if not live:
            for lock in _glob.glob(os.path.join(ver_dir, ".pzr*.tmp")):
                try:
                    os.remove(lock)
                    removed_locks += 1
                except OSError:
                    pass
        # Per-PID backup/run dirs — remove only dead PIDs
        run_dir = os.path.join(ver_dir, "run")
        if not os.path.isdir(run_dir):
            continue
        for pid_name in sorted(os.listdir(run_dir)):
            try:
                pid = int(pid_name)
            except ValueError:
                continue
            if pid in live:
                continue  # active Prism session — leave it alone
            shutil.rmtree(os.path.join(run_dir, pid_name), ignore_errors=True)
            removed_dirs += 1
    if removed_locks or removed_dirs:
        print(f"  Cleaned Prism recovery leftovers "
              f"({removed_dirs} run dirs, {removed_locks} lock files)")


def quit_prism():
    """Quit Prism cleanly so it does NOT leave autosave/recovery files behind.

    Prism marks the XML-written .pzfx as modified, so a plain 'quit' pops a
    'Save changes?' dialog that blocks shutdown (and force-killing it leaves
    recovery files that reopen as 'not well closed' documents next launch).
    We ask Prism to quit WITHOUT saving (results are already exported to
    Bead*.txt by the .pzc script), wait for the process to actually exit, and
    only fall back to pkill if it is genuinely stuck — then clean up the
    crash-recovery artifacts so the next launch starts fresh.

    NOTE: 'quit saving no' discards unsaved changes in Prism. The pipeline
    only opens its own temp files, whose results are already on disk, so
    nothing is lost for automation runs.
    """
    if sys.platform == "darwin":
        # 1) Graceful quit — don't save (exports are already on disk)
        try:
            subprocess.run(["osascript", "-e",
                            f'tell application "{PRISM_APP}" to quit saving no'],
                           capture_output=True, timeout=10)
        except subprocess.TimeoutExpired:
            cp.warn("  WARNING: osascript quit timed out")
        # 2) Wait up to 20 s for Prism to actually exit
        for _ in range(20):
            time.sleep(1)
            if not _prism_pids():
                break
        else:
            # 3) Plain quit attempt (some Prism versions don't accept 'saving no')
            print("  Prism still running — trying plain quit")
            try:
                subprocess.run(["osascript", "-e",
                                f'tell application "{PRISM_APP}" to quit'],
                               capture_output=True, timeout=10)
            except subprocess.TimeoutExpired:
                pass
            for _ in range(10):
                time.sleep(1)
                if not _prism_pids():
                    break
            else:
                cp.warn("  WARNING: Prism still running after graceful quit — force-killing")
                try:
                    subprocess.run(["pkill", "-f", PRISM_APP],
                                   capture_output=True, timeout=5)
                except Exception:
                    pass
                time.sleep(2)
    elif sys.platform == "win32":
        # Graceful close first, then force as fallback
        subprocess.run(["taskkill", "/im", "prism.exe"],
                       capture_output=True, timeout=5)
        time.sleep(3)
        subprocess.run(["taskkill", "/f", "/im", "prism.exe"],
                       capture_output=True, timeout=5)
        time.sleep(1)
    # Belt & braces: clear any leftover recovery state from a previous kill
    cleanup_prism_recovery()


def open_prism_and_save(pzfx_path, prism_path, beads_to_export=None):
    """Launch Prism headless, poll export script until analysis completes.

    Launches Prism hidden (no window focus), then runs the export script in
    a polling loop. Each cycle exports all 10 beads and checks Bead1 for
    valid parameters. Exits as soon as results appear. Prism stays open
    afterward for manual inspection if needed.
    """
    print(f"\n  Opening Prism...")
    script_template = os.path.join(PRISM_SCRIPT_DIR, "Automated datcol.pzc")
    script_path = os.path.join(PRISM_SCRIPT_DIR, "_auto_run_temp.pzc")
    bead1_path = os.path.join(PRISM_SCRIPT_DIR, "Bead1.txt")

    if not os.path.exists(script_template):
        raise FileNotFoundError(f"Prism script template not found: {script_template}")

    with open(script_template, "r", encoding="utf-8") as f:
        pzc_content = f.read()

    if "DONE_SIGNAL_PLACEHOLDER" not in pzc_content:
        cp.warn("  WARNING: Prism script template has no DONE_SIGNAL_PLACEHOLDER")

    done_path = done_file.replace("\\", "\\\\") if sys.platform == "win32" else done_file
    pzc_content = pzc_content.replace("DONE_SIGNAL_PLACEHOLDER", done_path)

    # If caller provided a list of beads to export, strip commands for missing beads
    if beads_to_export is not None:
        allowed = set(beads_to_export)
        lines = pzc_content.splitlines()
        out_lines = []
        for line in lines:
            stripped = line.strip()
            m_goto = re.match(r"GoTo\s+R(\d+)", stripped)
            m_export = re.match(r"ExportTable\s+Bead(\d+)\.txt", stripped)
            if m_goto:
                bead_name = f"Bead{m_goto.group(1)}"
                if bead_name not in allowed:
                    continue
            if m_export:
                bead_name = f"Bead{m_export.group(1)}"
                if bead_name not in allowed:
                    continue
            out_lines.append(line)
        pzc_content = "\n".join(out_lines)

    with open(script_path, "w", encoding="utf-8") as f:
        f.write(pzc_content)

    print(f"  Launching Prism file: {pzfx_path}")
    print(f"  pzfx exists: {os.path.exists(pzfx_path)}")
    if os.path.exists(pzfx_path):
        print(f"  pzfx size: {os.path.getsize(pzfx_path)} bytes")
    print(f"  script template: {script_template}")
    print(f"  generated script: {script_path}")
    print(f"  done file: {done_file}")

    # ── Kill any previous Prism instance to avoid stale data ──
    quit_prism()

    # ── Clean up old export files from previous runs ──
    import glob
    for f in glob.glob(os.path.join(PRISM_SCRIPT_DIR, "Bead*.txt")):
        try:
            os.remove(f)
        except OSError:
            pass
    if os.path.exists(done_file):
        try:
            os.remove(done_file)
        except OSError:
            pass

    if sys.platform == "darwin":
        if PRISM_HEADLESS:
            subprocess.run(["open", "-g", "-a", PRISM_APP, pzfx_path])
        else:
            subprocess.run(["open", "-a", PRISM_APP, pzfx_path])
    elif sys.platform == "win32":
        try:
            os.startfile(pzfx_path)
        except OSError:
            if PRISM_APP_PATH and os.path.exists(PRISM_APP_PATH):
                subprocess.Popen([PRISM_APP_PATH, pzfx_path])
            else:
                raise
    else:
        try:
            subprocess.run(["xdg-open", pzfx_path])
        except Exception:
            print("  Could not open Prism file automatically on this platform")

    # Wait for Prism to launch
    mode = "headless" if PRISM_HEADLESS else "foreground"
    # Poll for the Prism process instead of sleeping a fixed time (saves
    # several seconds per batch when Prism launches faster).
    print(f"  Waiting for Prism to launch ({mode})...")
    waited = 0
    while waited < PRISM_LOAD_DELAY and not _prism_pids():
        time.sleep(1)
        waited += 1
    time.sleep(3)  # let the document open and auto-analysis start

    # Hide/minimize Prism after launch (best-effort, may fail silently)
    if PRISM_HEADLESS:
        if sys.platform == "darwin":
            try:
                hide_script = f'''
                tell application "System Events"
                    set visible of process "{PRISM_APP}" to false
                end tell
                '''
                subprocess.run(["osascript", "-e", hide_script],
                               capture_output=True, timeout=5)
                print("  Prism hidden (Cmd+Tab to bring back).")
            except Exception:
                pass  # open -g already launched in background
        elif sys.platform == "win32":
            try:
                import ctypes
                hwnd = ctypes.windll.user32.FindWindowW(None, "Prism 11")
                if hwnd:
                    ctypes.windll.user32.ShowWindow(hwnd, 6)  # SW_MINIMIZE = 6
                    print("  Prism minimized.")
            except Exception:
                print("  Could not minimize Prism window.")

    # Polling: run Prism script -> check Bead1 export -> repeat until valid
    print(f"  Polling for analysis completion (every {PRISM_POLL_INTERVAL}s, "
          f"timeout {PRISM_POLL_TIMEOUT}s)...")
    start = time.time()
    attempt = 0

    while time.time() - start < PRISM_POLL_TIMEOUT:
        attempt += 1
        elapsed = int(time.time() - start)

        if os.path.exists(done_file):
            os.remove(done_file)
        if os.path.exists(bead1_path):
            os.remove(bead1_path)

        # Run the Prism export script
        if sys.platform == "darwin":
            applescript = f'''
            tell application "{PRISM_APP}"
                open POSIX file "{script_path}"
            end tell
            '''
            subprocess.run(["osascript", "-e", applescript])
        elif sys.platform == "win32":
            # On Windows, opening a .pzc file runs it in Prism
            os.startfile(script_path)
        else:
            subprocess.run(["xdg-open", script_path])

        # Wait for script to finish (done.txt appears)
        script_start = time.time()
        while not os.path.exists(done_file):
            if time.time() - script_start > 60:
                break
            time.sleep(1)

        # Check exports: accept ANY bead file with valid params.
        # (Bead1.txt is sometimes empty when the first results sheet has no
        # exportable data, while Bead2..Bead10 are fine — don't gate on Bead1.)
        import glob as _glob
        parsed_any = {}
        found_beads = []
        nonempty = 0
        for bf in sorted(_glob.glob(os.path.join(PRISM_SCRIPT_DIR, "Bead*.txt"))):
            p = parse_prism_export_txt(bf)
            if p:
                nonempty += 1
            if p and len(p) >= 2:
                found_beads.append(os.path.basename(bf))
                parsed_any.update(p)
        if found_beads:
            print(f"  Analysis complete after {elapsed}s (attempt {attempt})")
            print(f"  Results found in: {found_beads}")
            print(f"  Params: {parsed_any}")
            quit_prism()  # close Prism after use, don't save
            return

        print(f"  [{elapsed}s] No results yet (got {len(parsed_any)} params, "
              f"{nonempty} non-empty bead files), retrying in {PRISM_POLL_INTERVAL}s...")
        time.sleep(PRISM_POLL_INTERVAL)

    quit_prism()  # clean up on timeout
    raise TimeoutError(
        f"Prism analysis did not complete within {PRISM_POLL_TIMEOUT}s. "
        f"Check that the template has auto-analyze enabled and Prism is not stuck."
    )


def run_prism_script(script_name):
    """Run a Prism script by name via the Scripts menu.
    Returns True if the script was triggered.
    """
    # Try the Scripts menu first (Prism 11 stores scripts here)
    script = f'''
    tell application "{PRISM_APP}" to activate
    delay 1
    tell application "System Events"
        tell process "{PRISM_APP}"
            try
                click menu item "{script_name}" of menu 1 of menu item "Scripts" of menu bar 1
                return "OK"
            on error
                try
                    click menu item "{script_name}" of menu "Scripts" of menu bar 1
                    return "OK"
                on error
                    return "NOT_FOUND"
                end try
            end try
        end tell
    end tell
    '''
    result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    return result.stdout.strip()


def parse_prism_export_txt(txt_path, wanted=None):
    """Parse a Prism-exported .txt file for best-fit values.
    
    Expected format (tab-delimited, quoted strings):
        "Best-fit values"
        "     k0"	0.3738
        "     k1"	0.6416
        "     tau"	9.533
        "     viscosity"	398.2
    
    Returns dict {param_name: value_string}.
    """
    if wanted is None:
        wanted = WANTED_PARAMS

    with open(txt_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    results = {}
    in_best_fit = False

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Detect section headers (no tab character)
        if '\t' not in line:
            if 'Best-fit' in line:
                in_best_fit = True
            elif in_best_fit:
                # Hit a new section — stop collecting
                break
            continue

        if not in_best_fit:
            continue

        # Parse: "     k0"\t0.3738
        parts = line.split('\t')
        if len(parts) < 2:
            continue

        param_name = parts[0].strip().strip('"').strip()
        value_str = parts[1].strip().strip('"')
        value_str = re.sub(r'^[~≈≃]+\s*', '', value_str).strip()

        if param_name in wanted:
            results[param_name] = value_str

    return results


def extract_prism_results(prism_path, table_name, wanted=None):
    """Extract best-fit values from a Prism .prism file (ZIP with JSON).
    Returns list of (param_name, value) tuples.
    """
    import zipfile
    import json

    if wanted is None:
        wanted = WANTED_PARAMS

    results = []

    with zipfile.ZipFile(prism_path, 'r') as z:
        analysis_dirs = set()
        for name in z.namelist():
            if name.startswith("analyses/") and name.endswith("/sheet.json"):
                analysis_dirs.add(name.rsplit("/", 1)[0])

        for analysis_dir in analysis_dirs:
            try:
                sheet_json = json.loads(z.read(f"{analysis_dir}/sheet.json"))
                title = sheet_json.get("title", "")
                if table_name not in title:
                    continue

                results_json = json.loads(z.read(f"{analysis_dir}/results.json"))
                models = results_json.get("content", {}).get("models", {})

                for model_id, model_data in models.items():
                    params = model_data.get("value", {}).get("parameters", {})
                    for param_name in wanted:
                        if param_name in params:
                            param_data = params[param_name]
                            for key, val_wrapper in param_data.items():
                                if key == "id":
                                    continue
                                estimate = val_wrapper.get("value", {}).get("estimate", "")
                                if estimate != "":
                                    results.append((param_name, str(round(estimate, 4))))
                                    print(f"    {param_name} = {round(estimate, 4)}")

            except (KeyError, json.JSONDecodeError):
                continue

    return results


# ══════════════════════════════════════════════════════════════
#  GROUPED PRISM BUILDER — creates .pzfx with G'(0.1/1/10 Hz)
#  tables grouped by condition, with per-sample subcolumns
# ══════════════════════════════════════════════════════════════


def discover_samples_for_batch(base_dir):
    """Find all sample folders and infer conditions from parent dirs.

    Walks base_dir for folders ending in ' separated' &'.separated' and uses the
    immediate parent folder name as the condition.

    Returns {condition: {sample_name: Path_to_sample_dir}}
    """
    base = Path(base_dir).expanduser().resolve()
    conditions = {}
    # Collect matches from both naming patterns, keep order deterministic,
    # and remove duplicates (some paths might match both patterns).
    from itertools import chain

    matches = list(base.rglob("* separated")) + list(base.rglob("*.separated"))
    # Deduplicate while preserving order
    seen = set()
    unique_matches = []
    for p in matches:
        try:
            key = p.resolve()
        except Exception:
            key = p
        if key in seen:
            continue
        seen.add(key)
        unique_matches.append(p)

    for sample_dir in sorted(unique_matches, key=natural_sort_key):
        if not sample_dir.is_dir():
            continue
        condition = sample_dir.parent.name
        sample_name = sample_dir.name.replace(" separated", "").replace(".separated", "")
        conditions.setdefault(condition, {})[sample_name] = sample_dir
    return conditions


def build_grouped_prism(data, times, output_path):
    """Create a .pzfx with grouped tables from a nested data dict.

    data = {
        "Gp_0.1Hz": {"AMI21": {"p1b2": [val,...], "p2b2": [...]}, "PD62": {...}},
        "Gp_1Hz":   {...},
        "Gp_10Hz":  {...},
    }
    times = [7, 14, 21, ...]
    """

    ET.register_namespace("", _NS)
    ET.register_namespace("dt", "urn:schemas-microsoft-com:datatypes")

    table_names = list(data.keys())
    if not table_names:
        raise ValueError("data dict is empty")

    first = data[table_names[0]]
    conditions = list(first.keys())
    sample_map = {c: list(first[c].keys()) for c in conditions}

    root = ET.Element(ns("GraphPadPrismFile"))
    root.set("PrismXMLVersion", "7.00")

    info = ET.SubElement(root, ns("Info"))
    info.set("Info table", "")
    info.set("Sequence", "1")
    info_table = ET.SubElement(info, ns("Table"))
    info_table.set("TableType", "Info")
    ET.SubElement(info_table, ns("Title")).text = "Info"

    sheet = ET.SubElement(root, ns("Sheet"))
    sheet.set("Sequence", "data")
    sheet.set("Name", "Data")

    for table_title, table_data in data.items():
        table_el = ET.SubElement(sheet, ns("Table"))
        table_el.set("TableType", "Grouped")
        table_el.set("XFormat", "none")
        ET.SubElement(table_el, ns("Title")).text = table_title

        # Row titles (time points)
        rt = ET.SubElement(table_el, ns("RowTitlesColumn"))
        rt.set("Width", "60")
        rt.set("Decimals", "0")
        rsub = ET.SubElement(rt, ns("Subcolumn"))
        for t in times:
            d = ET.SubElement(rsub, ns("d"))
            d.text = str(t)

        # Y columns — one per condition, subcolumns per sample
        for cond in conditions:
            samples = sample_map[cond]
            yc = ET.SubElement(table_el, ns("YColumn"))
            yc.set("Width", "81")
            yc.set("Subcolumns", str(len(samples)))
            yc.set("Decimals", "4")
            ET.SubElement(yc, ns("Title")).text = cond

            for sname in samples:
                sc = ET.SubElement(yc, ns("Subcolumn"))
                ET.SubElement(sc, ns("Title")).text = sname
                vals = table_data[cond][sname]
                if not isinstance(vals, (list, tuple)):
                    vals = [vals] * len(times)
                for v in vals:
                    d = ET.SubElement(sc, ns("d"))
                    d.text = str(v) if v is not None else ""

    # Write with Prism-compatible formatting
    tree = ET.ElementTree(root)
    tree.write(output_path, xml_declaration=True, encoding="utf-8")

    with open(output_path, "r", encoding="utf-8") as f:
        content = f.read()
    content = content.replace(
        "<?xml version='1.0' encoding='utf-8'?>",
        '<?xml version="1.0" encoding="UTF-8"?>'
    )
    content = content.replace('xmlns:ns1=', 'xmlns:dt=')
    content = content.replace('ns1:', 'dt:')
    content = content.replace("\n", "\r\n")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"\nGrouped Prism saved → {output_path}")
    print(f"  Tables: {table_names}")
    print(f"  Conditions: {conditions}")
    for c, s in sample_map.items():
        print(f"    {c}: {s}")

    return output_path




# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    # ── BATCH MODE (orchestrator) ──
    if BATCH_MODE and not os.environ.get("PRISM_SINGLE_RUN"):
        import subprocess as sp

        print("=" * 50)
        cp.ok("BATCH MODE — processing all samples for grouped G' tables")
        print("=" * 50)

        base = Path(BASE_DIR_FOR_BATCH).expanduser()
        conditions = discover_samples_for_batch(base)

        if not conditions:
            cp.err("No sample folders found.")
            sys.exit(1)

        total = sum(len(s) for s in conditions.values())
        print(f"\nFound {total} samples across {len(conditions)} conditions:\n")
        for cond, samples in sorted(conditions.items()):
            print(f"  {cond}/")
            for name in sorted(samples):
                print(f"    {name}")

        script_path = Path(__file__).resolve()
        python_exe = sys.executable
        done = 0

        skipped_samples = {}
        missing_outputs = []   # per-sample list; reset each iteration below
        for condition, samples in sorted(conditions.items()):
            cond_dir = base / condition
            for sample_name, sample_dir in sorted(samples.items()):
                done += 1
                sample_path = str(sample_dir)

                # Skip if an output file for this sample already exists
                output_dir = sample_dir.parent
                pattern = f"{condition}_{sample_name}_bead*.xlsx"
                candidates = sorted(output_dir.glob(pattern))
                candidates = [p for p in candidates if not p.name.startswith("~$")]



                sorted_samples = sorted(sample_dir.glob("*.txt"), key=natural_sort_key)
                filtered_samples = [f for f in sorted_samples if "pulling" in f.name and os.path.getsize(f) > 93000]

                batch_size = BEAD_COUNT
                num_batches = len(filtered_samples) // batch_size
                remainder = len(filtered_samples) % batch_size
                first_num = get_number(filtered_samples[0].name) if filtered_samples else None

                if remainder > 0:
                    cp.warn(f"    WARNING: sample has {len(filtered_samples)} pulling files, which is not a multiple of {batch_size}. "
                          f"Only {num_batches} full batches will be checked.")


                if num_batches <= 0 or first_num is None:
                    # No valid pulling files: if an output already exists treat
                    # the sample as done; otherwise it gets reported below.
                    processed = len(candidates) > 0
                    expected_outputs = []
                    expected_names = set()
                    actual_names = {p.name for p in candidates}
                    missing_outputs = []
                else:
                    expected_outputs = []
                    missing_outputs = []
                    # Build the batches from the ACTUAL valid file list, exactly
                    # like the single-run pipeline does (batched(filtered,
                    # BEAD_COUNT, strict=False), skipping incomplete tails).
                    # The expected output name must be bead{first}-{tenth} of
                    # that real batch — NOT bead{first}-{first+9}. If a pulling
                    # file is missing from the middle (e.g. below the size
                    # cutoff), the batch's 10th file has a different number,
                    # and an arithmetically computed name would never match
                    # what the pipeline produces -> the sample would look
                    # 'missing' forever and re-run in a loop on every run.
                    for batch_files in batched(filtered_samples, batch_size, strict=False):
                        if len(batch_files) != batch_size:
                            continue  # incomplete tail batch — never processed
                        start = get_number(batch_files[0].name)
                        end = get_number(batch_files[batch_size - 1].name)
                        if start is None or end is None:
                            continue
                        expected_path = output_dir / f"{condition}_{sample_name}_bead{start}-{end}.xlsx"
                        expected_outputs.append(expected_path)
                        if not expected_path.exists():
                            missing_outputs.append(expected_path)
                    expected_names = {p.name for p in expected_outputs}
                    actual_names = {p.name for p in candidates}
                    # All expected batch outputs exist -> sample is processed.
                    # Do NOT also require len(candidates) == num_batches: a
                    # stale/extra output file (old run, changed numbering) would
                    # otherwise make the sample look 'unprocessed' forever and
                    # the batch would re-run it on every invocation — an
                    # endless loop across runs that piles up more files.
                    processed = len(missing_outputs) == 0

                if not processed:
                    print(f"  [check] {condition}/{sample_name}: found {len(candidates)} output files, expected {num_batches} batches")
                    print(f"Checking individual batches for {condition}/{sample_name}...")
                    if first_num is None:
                        cp.warn("    No valid pulling files found for this sample.")
                    else:
                        for batch_idx, expected_path in enumerate(expected_outputs):
                            if not expected_path.exists():
                                print(f"    Missing expected output for batch {batch_idx}: {expected_path.name}")
                        extra_files = sorted(actual_names - expected_names)
                        if extra_files:
                            print(f"    Unexpected output files: {', '.join(extra_files)}")

                print(f"  [check] {condition}/{sample_name}: processed={processed}")
                if processed:
                    cp.info(f"    Found existing output: {', '.join(p.name for p in candidates)}")
                    cp.ok(f"\n[{done}/{total}] {condition}/{sample_name} — already done, skipping")
                    continue

                print(f"\n{'='*50}")
                print(f"[{done}/{total}] Running pipeline for {condition}/{sample_name}")
                print(f"  Folder: {sample_path}") 
                print(f"{'='*50}")

                env = os.environ.copy()
                env["PRISM_SAMPLE_DIR"] = sample_path
                env["PRISM_SINGLE_RUN"] = "1"

                # If we detected specific missing batch files, run only those batches
                if missing_outputs:
                    for expected_path in missing_outputs:
                        # expected filename like '{condition}_{sample_name}_bead{start}-{end}.xlsx'
                        m = re.search(r"bead(\d+)-(\d+)\.xlsx$", expected_path.name)
                        if not m:
                            continue
                        start = int(m.group(1))
                        end = int(m.group(2))
                        env_batch = env.copy()
                        env_batch["PRISM_BATCH_START"] = str(start)
                        env_batch["PRISM_BATCH_END"] = str(end)
                        cp.info(f"  Running missing batch {start}-{end} for {condition}/{sample_name}...")
                        try:
                            result = subprocess.run([python_exe, str(script_path)],
                                            capture_output=False, timeout=1200, env=env_batch)
                        except subprocess.TimeoutExpired:
                            cp.warn(f"  WARNING: batch {start}-{end} timed out after 1200s — "
                                  f"it will be retried on the next run")
                            continue
                        if result.returncode != 0:
                            cp.warn(f"  WARNING: batch {start}-{end} exited with code {result.returncode}")
                        else:
                            cp.ok(f"  Batch {start}-{end} complete for {condition}/{sample_name}")
                else:
                    # No specific missing batches detected — run whole-sample pipeline
                    cp.info(f"  Running whole-sample pipeline...")
                    try:
                        result = subprocess.run([python_exe, str(script_path)],
                                        capture_output=False, timeout=1200, env=env)
                    except subprocess.TimeoutExpired:
                        cp.warn(f"  WARNING: pipeline timed out after 1200s — "
                              f"it will be retried on the next run")
                        continue
                    if result.returncode != 0:
                        cp.warn(f"  WARNING: pipeline exited with code {result.returncode}")
                    else:
                        cp.ok(f"  Pipeline complete for {condition}/{sample_name}")

                report_path = sample_dir / "missing_beads.txt"
                if report_path.exists():
                    with open(report_path, "r", encoding="utf-8") as f:
                        missing = [line.strip() for line in f.readlines() if line.strip() and not line.startswith("Missing beads")]
                    skipped_samples[f"{condition}/{sample_name}"] = missing

        if skipped_samples:
            print("\nBatch summary: samples skipped due to missing beads:")
            for sample_id, missing in skipped_samples.items():
                print(f"  {sample_id}: missing {len(missing)} beads -> {', '.join(missing)}")

        cp.ok("\nBatch pipeline complete. G' values in Excel — run final_result.py to build grouped Prism.")
        sys.exit(0)

    # ── Single-sample pipeline (or subprocess spawned by batch mode) ──
    # Allow batch mode to override FOLDER_PATH via env var
    _batch_folder = os.environ.get("PRISM_SAMPLE_DIR")
    if _batch_folder:
        FOLDER_PATH = _batch_folder
        print(f"  [batch] FOLDER_PATH={FOLDER_PATH}")

    folder_path = Path(FOLDER_PATH)

    try:
        sorted_files = sorted(folder_path.glob("*.txt"), key=natural_sort_key)
        filtered = [f for f in sorted_files if "pulling" in f.name and os.path.getsize(f) > 93000]

        # Allow orchestrator to request a single batch via env vars
        batch_only_start = None
        env_bs = os.environ.get("PRISM_BATCH_START")
        if env_bs:
            try:
                batch_only_start = int(env_bs)
                print(f"  [single-run] Requested to run only batch starting at {batch_only_start}")
            except Exception:
                batch_only_start = None

        batch = batched(filtered, BEAD_COUNT, strict=False)
        t = 0
        t_batch = time.time()   # timing: whole sample (all batches) processing

        for filtered_files in batch:
            # Skip an incomplete tail batch (e.g. 23 files -> batches of 10 + 3).
            # strict=False avoids the ValueError that strict=True would raise
            # mid-iteration and crash the whole sample.
            if len(filtered_files) != BEAD_COUNT:
                cp.warn(f"  Skipping incomplete tail batch ({len(filtered_files)} files) — "
                      f"only full batches of {BEAD_COUNT} are processed.")
                continue
            # If orchestrator requested a specific batch, skip others
            if batch_only_start is not None:
                this_start = get_number(filtered_files[0].name)
                if this_start != batch_only_start:
                    continue
            t += 1
            active_bead_count = BEAD_COUNT   # lowered when trailing beads are missing
            # print(f"Running batch {t} of {os.path.basename(folder_path)} with {len(filtered_files)} files...")

            print(f"Found {len(filtered_files)} data files\n")

            if len(filtered_files) < 1:
                cp.err("No data files found!")
                sys.exit(1)

            elif len(filtered_files) < (BEAD_COUNT - 2):
                cp.warn(f"⚠️ WARNING: Found only {len(filtered_files)} data files, expected {BEAD_COUNT}.")
                cp.err("❌☠️ Not proceeding with available files. Prism export will fail for missing beads.")
                sys.exit(1)

            elif len(filtered_files) < (BEAD_COUNT- 1):
                cp.warn(f"⚠️ WARNING: Found only {len(filtered_files)} data files, expected {BEAD_COUNT}.")
                print("Proceeding with available files. Prism export may fail for missing beads but we move.")

            elif len(filtered_files) > 10:
                cp.info("Found more than 10 files, checking if more pulling data can be extracted")
                
        
            first_num = get_number(filtered_files[0].name)
            last_idx = min(BEAD_COUNT - 1, len(filtered_files) - 1)
            tenth_num = get_number(filtered_files[last_idx].name)

            # print("There are ", math.floor(filtered_files/10), "x 10 pulling data")

            # if math.floor(filtered_files/10) == 1: 
            #     first_num = get_number(filtered_files[0].name)
            #     last_idx = min(BEAD_COUNT - 1, len(filtered_files) - 1)
            #     tenth_num = get_number(filtered_files[last_idx].name)
            
            # elif math.floor(filtered_files/10) > 1:
            #     iterator = iter(filtered_files)

            #     while math.floor(filtered_files/10) > 1 == True: 
            #         batch = list(islice(iterator, 10))

                

                






            # ── Step 1: Read .txt data files ──
            cp.section("Step 1: Reading data files...")
            dataframes = []
            for f in filtered_files[:BEAD_COUNT]:
                df = pd.read_csv(f, sep="\t")
                dataframes.append(df)
                print(f"  {f.name}: {len(df)} rows, {len(df.columns)} cols")

            # ── Step 2: Write data to Excel and read computed values ──
            cp.section("\nStep 2: Opening Excel with xlwings...")
            template_path = Path(EXCEL_TEMPLATE)
            if not template_path.exists():
                cp.err(f"ERROR: Excel template not found: {template_path}")
                sys.exit(1)

            app = xw.App(visible=True, add_book=False)
            try:
                wb_xl = app.books.open(str(template_path))
            except Exception as exc:
                cp.err(f"ERROR: Unable to open Excel template: {template_path}")
                print(f"  {type(exc).__name__}: {exc}")
                app.quit()
                sys.exit(1)

            ws_input = wb_xl.sheets[INPUT_SHEET]

            # Write .txt data to "Data Input from Text File" sheet (bulk)
            col_i = 1
            for df in dataframes:
                data = df.values.tolist()
                start_cell = ws_input.cells(2, col_i)
                start_cell.resize(len(data), len(data[0])).value = data
                col_i += 6
            cp.info("  Wrote data to Input sheet.")

            # Save so Excel computes formulas
            excel_output = build_excel_output_path(folder_path, f"{first_num}-{tenth_num}")
            wb_xl.save(str(excel_output))
            time.sleep(2)  # Let Excel finish computing
            print(f"  Saved: {excel_output}")

            

            # Read computed values from each Bead sheet (columns T=20, U=21)
            bead_data = {}
            for j in range(BEAD_COUNT):
                bead_name = f"Bead{j + 1}"
                print(f"\n  Processing {bead_name}...")
                try:
                    ws_bead = wb_xl.sheets[bead_name]
                except Exception as e:
                    print(f"  [{bead_name}] Sheet not found: {e}")
                    continue

                # Check K1 for 'OK' — temporarily disabled for debugging
                k1_val = ws_bead.range("K1").value
                print(f"  [{bead_name}] K1={k1_val}")
                # if ws_bead.range("K1").value == 'OK':
                #     k1_val = ws_bead.range("K1").value
                #     print(f"  [{bead_name}] checked — marked OK -- [{k1_val}]")
                #     continue

                # Check if data exists
                a4_val = ws_bead.range("A4").value
                print(f"  [{bead_name}] A4={a4_val}")
                if str(a4_val) == '0' or a4_val is None:
                    print(f"  [{bead_name}] Data sheet empty, skipping.")
                    continue

                # Read columns T and U (starting at row 2 to skip header)
                t1_val = ws_bead.range("T2").value
                u1_val = ws_bead.range("U2").value
                print(f"  [{bead_name}] T2={t1_val}, U2={u1_val}")

                col_t = ws_bead.range("T2").expand("down").value
                col_u = ws_bead.range("U2").expand("down").value

                # Handle single value (not a list)
                if not isinstance(col_t, list):
                    col_t = [col_t]
                if not isinstance(col_u, list):
                    col_u = [col_u]

                # Make equal length
                min_len = min(len(col_t), len(col_u))
                col_t = col_t[:min_len]
                col_u = col_u[:min_len]

                xy = pd.DataFrame({"X": col_t, "Y": col_u})
                print(xy)
                print(f"  [{bead_name}] T/U: {len(xy)} rows (first: X={col_t[0]}, Y={col_u[0]})")

                if len(xy) > 0:
                    bead_data[bead_name] = xy

            # Save and close Excel
            wb_xl.save()
            wb_xl.close()
            app.kill()
            cp.ok(f"\n  Excel closed. Beads with data: {list(bead_data.keys())}")

            # Check for missing beads and inform user
            all_beads = [f"Bead{i+1}" for i in range(BEAD_COUNT)]
            beads_with_data_list = list(bead_data.keys())
            missing_beads = [b for b in all_beads if b not in beads_with_data_list]

            marker_path = Path(folder_path) / "missing_beads.txt"
            if missing_beads:
                # If the missing beads are only at the END (trailing, e.g. just
                # bead 10), continue with the beads we have and record nil for
                # the missing ones. Otherwise skip the sample as before.
                present_nums = sorted(int(b.replace("Bead", "")) for b in beads_with_data_list)
                trailing_missing = bool(present_nums) and present_nums == list(range(1, len(present_nums) + 1))
                if trailing_missing:
                    active_bead_count = len(present_nums)
                    cp.warn(f"\n  ⚠️  WARNING: Only {active_bead_count}/{BEAD_COUNT} beads have data "
                            f"(missing: {', '.join(missing_beads)}) — continuing with {active_bead_count} beads; "
                            f"G' for the missing trailing bead(s) will be nil.")
                else:
                    cp.warn(f"\n  ⚠️  WARNING: The following beads have no data — skipping this sample:")
                    for bead in missing_beads:
                        print(f"      - {bead}")
                    print(f"  Found {len(beads_with_data_list)} beads with data; need {BEAD_COUNT}.")

                    with open(marker_path, "w", encoding="utf-8") as f:
                        f.write("Missing beads:\n")
                        for bead in missing_beads:
                            f.write(f"{bead}\n")

                    cp.warn(f"  Skipping Prism/MATLAB processing for sample. Marker written: {marker_path}")
                    sys.exit(0)
            else:
                if marker_path.exists():
                    marker_path.unlink()
                cp.ok(f"\n  ✓ All {BEAD_COUNT} beads have data.")

            # ── Step 3: Copy template and write data via XML ──
            cp.section("\nStep 3: Writing data to Prism template...")
            print(f"Platform: {'Windows' if is_windows() else 'macOS' if is_macos() else sys.platform}")
            # prism_dir = os.path.expanduser("~/prism_automation")
            # os.makedirs(prism_dir, exist_ok=True)
            # temp_pzfx = os.path.join(prism_dir, f"prism_temp_{first_num}-{tenth_num}.pzfx")
            prism_dir = excel_output.parent
            os.makedirs(prism_dir, exist_ok=True)
            temp_pzfx = excel_output.with_suffix(".pzfx")
            # prism_file = temp_pzfx.replace(".pzfx", ".prism")
            prism_file = temp_pzfx.with_suffix(".prism")
            shutil.copy2(PZFX_TEMPLATE, temp_pzfx)
            print(f"  Copied template to: {temp_pzfx}")

            tree = ET.parse(temp_pzfx)
            root = tree.getroot()

            for j in range(BEAD_COUNT):
                bead_name = f"Bead{j + 1}"
                if bead_name not in bead_data:
                    continue

                xy = bead_data[bead_name]
                write_data_to_prism(root, bead_name, xy)
                ok, msg = validate_prism_table_columns(root, bead_name)
                if not ok:
                    cp.warn(f"  WARNING: [{bead_name}] Prism validation failed: {msg}")
                else:
                    print(f"  [{bead_name}] Prism validation OK")

            save_pzfx(tree, temp_pzfx)
            print(f"  Saved .pzfx with all bead data")

            # ── Step 4: Open Prism for analysis ──
            cp.section("\nStep 4: Opening Prism for analysis...")
            # Pass only beads that have data so the Prism script skips empty beads
            open_prism_and_save(temp_pzfx, prism_file, beads_to_export=beads_with_data_list)

            # ── Step 5: Extract results (try pzfx XML first, fallback to Prism script) ──
            # ── Step 5: Parse exported .txt files into Excel ──
            cp.section("\nStep 5: Parsing exported results into Excel...")
            wb_xl = None   # Excel session — opened in Step 6, reused in Step 10
            app_xl = None
            wb = load_workbook(excel_output)
            script_dir = PRISM_SCRIPT_DIR
            
            # Track which beads have valid data
            beads_with_data = []

            for j in range(BEAD_COUNT):
                bead_name = f"Bead{j + 1}"
                if bead_name not in wb.sheetnames:
                    continue

                ws = wb[bead_name]
                txt_path = os.path.join(script_dir, f"{bead_name}.txt")
                if not os.path.exists(txt_path):
                    txt_path = os.path.join(script_dir, bead_name)

                if not os.path.exists(txt_path):
                    print(f"  [{bead_name}] Export file not found: {txt_path}")
                    ws.sheet_properties.tabColor = "FF8C00"
                    continue

                parsed = parse_prism_export_txt(txt_path)
                if parsed:
                    has_issue = False
                    for i, param in enumerate(WANTED_PARAMS):
                        if param in parsed:
                            val_str = parsed[param]
                            try:
                                val_float = float(val_str)
                                ws.cell(row=2 + i, column=17, value=val_float)
                            except ValueError:
                                ws.cell(row=2 + i, column=17, value=val_str)
                                has_issue = True
                                cp.warn(f"  [{bead_name}] WARNING: {param} = '{val_str}' (not a number)")

                            if param == "viscosity":
                                if 'e' in val_str.lower():
                                    has_issue = True
                                    cp.warn(f"  [{bead_name}] WARNING: viscosity is very large ({val_str})")
                                try:
                                    if float(val_str) > 10000:
                                        has_issue = True
                                        cp.warn(f"  [{bead_name}] WARNING: viscosity = {val_str} (very large)")
                                except ValueError:
                                    pass

                    missing = [p for p in WANTED_PARAMS if p not in parsed]
                    if missing:
                        has_issue = True
                        cp.warn(f"  [{bead_name}] WARNING: missing params: {missing}")

                    if has_issue:
                        ws.sheet_properties.tabColor = "FF8C00"
                        print(f"  [{bead_name}] Tab colored orange (issues detected)")
                    # else:
                        # Track beads with valid data (no issues)
                    
                    beads_with_data.append(j + 1)
                    print(f"  [{bead_name}] Wrote {len(parsed)} values: {parsed}")
                else:
                    ws.sheet_properties.tabColor = "FF8C00"
                    print(f"  [{bead_name}] No best-fit values found — tab colored orange")
            
            # Color empty/unused bead sheets purple
            print(f"\n  Beads with valid data: {beads_with_data}")
            if beads_with_data:
                max_bead_with_data = max(beads_with_data)
                for j in range(max_bead_with_data, BEAD_COUNT):
                    bead_name = f"Bead{j + 1}"
                    if bead_name in wb.sheetnames:
                        ws = wb[bead_name]
                        ws.sheet_properties.tabColor = "800080"  # Purple
                        print(f"  [{bead_name}] Tab colored purple (unused/empty)")

            wb.save(excel_output)
            print(f"  Results saved to: {excel_output}")

            # ── Bead-count gate (after Prism) ──────────────────────────────
            # Fewer than BEAD_COUNT beads with results after Prism breaks
            # everything downstream: partial "Data Output to Matlab" sheets
            # give file2 an odd/partial column count -> MATLAB zeros() crash,
            # or silently wrong G' values. Stop here instead.
            missing_after_prism = [f"Bead{i + 1}" for i in range(BEAD_COUNT)
                                   if (i + 1) not in beads_with_data]
            if missing_after_prism:
                # Trailing-missing (e.g. just bead 10): continue with the beads
                # that have results; G' for the missing trailing bead(s) -> nil.
                present_nums = sorted(beads_with_data)
                trailing_missing = bool(present_nums) and present_nums == list(range(1, len(present_nums) + 1))
                if trailing_missing:
                    active_bead_count = len(present_nums)
                    cp.warn(f"\n  ⚠️  Only {active_bead_count}/{BEAD_COUNT} beads have results after Prism "
                            f"(missing: {', '.join(missing_after_prism)}) — continuing with {active_bead_count} beads; "
                            f"G' for the missing trailing bead(s) will be nil.")
                else:
                    cp.warn(f"\n  ⚠️  Only {len(beads_with_data)}/{BEAD_COUNT} beads have results "
                            f"after Prism (missing: {', '.join(missing_after_prism)}).")
                    cp.warn("  Skipping trim/MATLAB/summary for this sample — nothing downstream will be polluted.")
                    with open(marker_path, "w", encoding="utf-8") as f:
                        f.write("Missing beads (after Prism):\n")
                        for b in missing_after_prism:
                            f.write(f"{b}\n")
                    sys.exit(0)

            cp.section("\nStep 6: Trimming Data Output to Matlab 1 to shortest column...")

            app_xl = None
            wb_xl = None

            try:
                app_xl = xw.App(visible=False)
                wb_xl = app_xl.books.open(str(excel_output))
                ws_xl = wb_xl.sheets["Data Output to Matlab 1"]

                used = ws_xl.used_range

                values = used.value

                if not values:
                    print("  Sheet empty.")
                else:

                    # Ensure 2D
                    if not isinstance(values[0], list):
                        values = [values]

                    last_row = len(values)
                    last_col = max(len(row) for row in values)

                    print(f"  Sheet size: {last_row} rows x {last_col} cols")

                    shortest_end = last_row

                    for col in range(last_col):

                        col_end = 0

                        for row in range(last_row - 1, -1, -1):

                            value = (
                                values[row][col]
                                if col < len(values[row])
                                else None
                            )

                            if value is None:
                                continue

                            if isinstance(value, str) and value.strip() == "":
                                continue

                            if value == 0:
                                continue

                            col_end = row + 1
                            break

                        if col_end > 0:
                            shortest_end = min(shortest_end, col_end)

                    print(f"  Shortest data column ends at row {shortest_end}")

                    if shortest_end < last_row:

                        clear_range = ws_xl.range(
                            (shortest_end + 1, 1),
                            (last_row, last_col)
                        )

                        clear_range.clear_contents()

                        print(
                            f"  Cleared rows "
                            f"{shortest_end + 1} to {last_row}"
                        )

                    else:
                        print("  No trimming needed.")

                wb_xl.save()

                cp.section("\nStep 7: Copying MATLAB input 1 & 2 to file 2 and 1 respectively...")

                data = ws_xl.used_range.value
                if active_bead_count < BEAD_COUNT:
                    data = trim_matlab_columns(data, active_bead_count)

                os.makedirs(os.path.dirname(file2_dir), exist_ok=True)
                write_matlab_input(file2_dir, data)

                ws_xl = wb_xl.sheets["Data Output to Matlab 2"]
                data = ws_xl.used_range.value
                if active_bead_count < BEAD_COUNT:
                    data = trim_matlab_columns(data, active_bead_count)
                write_matlab_input(file1_dir, data)

                print("Suceesfully wrote to file1 and file2")   
                


            finally:
                # Keep Excel open so Step 10 can reuse this session — opening
                # + recalculating the 27MB workbook is the slowest part of a
                # batch. Only close here if Step 6/7 itself failed.
                if sys.exc_info()[0] is not None:
                    if wb_xl is not None:
                        wb_xl.close()
                    if app_xl is not None:
                        app_xl.quit()

            print(f"  Saved trimmed workbook to: {excel_output}")

    # ══════════════════════════════════════════════════════════════
    #  MATLAB
    # ══════════════════════════════════════════════════════════════

            cp.section("\nStep 8: Running MATLAB...")

            if sys.platform == "win32":
                analysis_dir = AUTOMATION_REQUIREMENTS_PATH
            else:
                analysis_dir = AUTOMATION_REQUIREMENTS_PATH
            os.makedirs(analysis_dir, exist_ok=True)
            matlab_script = os.path.join(analysis_dir, "moduliloop4_edited.m")
            file1 = os.path.join(analysis_dir, "file1.txt")
            file2 = os.path.join(analysis_dir, "file2.txt")
            csv_path = os.path.join(analysis_dir, "AllBeads.csv")
            print(f"  MATLAB script: {matlab_script}")
            print(f"  file1: {file1}")
            print(f"  file2: {file2}")
            print(f"  csv_path: {csv_path}")

            if sys.platform == "win32":
                # MATLAB on Windows needs paths with forward slashes or escaped backslashes
                matlab_script_escaped = matlab_script.replace("\\", "/")
                file1_escaped = file1.replace("\\", "/")
                file2_escaped = file2.replace("\\", "/")
            else:
                matlab_script_escaped = matlab_script
                file1_escaped = file1
                file2_escaped = file2

            cmd = (
                f"file1_path='{file1_escaped}';"
                f"file2_path='{file2_escaped}';"
                f"run('{matlab_script_escaped}');"
            )

            # A failed MATLAB run must never reuse a stale AllBeads.csv from a
            # previous batch — delete it first, then check the exit code so a
            # failed batch is reported and skipped, not silently over-written.
            if os.path.exists(csv_path):
                os.remove(csv_path)
                print(f"  Removed stale {os.path.basename(csv_path)} before this batch")
            result = subprocess.run(
                [MATLAB_CMD, "-batch", cmd],
                check=False
            )
            if result.returncode != 0:
                cp.warn(f"  WARNING: MATLAB exited with code {result.returncode} — "
                        f"{os.path.basename(csv_path)} not produced for this batch.")
            else:
                cp.ok("MATLAB has finished...")

            cp.section("\nStep 9: Processing AllBeads output...")
            
            df = None
            if os.path.exists(csv_path):
                df = pd.read_csv(csv_path, header=None)
                print(f"  Loaded {csv_path}: {df.shape[0]} rows x {df.shape[1]} cols")
                print(df.head())

                # Save a per-sample copy so batch_gp_prism can find it later
                # sample_copy = folder_path.parent / f"AllBeads_{folder_path.name.replace(' separated', '')}.csv"
                # df.to_csv(sample_copy, index=False, header=False)
                # print(f"  Saved per-sample copy: {sample_copy}")
            else:
                cp.warn(f"  WARNING: {csv_path} not found — MATLAB may not have produced output.")
            
            cp.section("\nStep 10: Inputting AllBeads into excel...")

            if df is None:
                cp.warn("  Skipping — no AllBeads data")
                # close the Excel session we kept open from Step 6
                if wb_xl is not None:
                    try:
                        wb_xl.close()
                    except Exception:
                        pass
                if app_xl is not None:
                    try:
                        app_xl.quit()
                    except Exception:
                        pass
            else:
                # reuse the Excel session kept open from Step 6 — no relaunch,
                # no 27MB reopen (this was one of the slowest steps)
                wb_cel = wb_xl
                app_cel = app_xl

                try:
                    ws_cel = wb_cel.sheets["Data Input from Matlab"]

                    data = df.values.tolist()
                    start_cell = ws_cel.cells(3, 1)
                    start_cell.value = data
                    print(f"  Wrote {len(data)} rows x {len(data[0]) if data else 0} cols to Data Input from Matlab")
        
                    wb_cel.save() 

                    
                    cp.section("\n Step 11: Outputting final results to Txt type file for final results...")
                    append_gp_and_gdp_summary(folder_path, excel_output, first_num, len(filtered_files),
                                              bead_count=active_bead_count)
                    # ws_out = wb_cel.sheets["For GraphPad Graph&Stats"]    

                    # pointone = ws_out.range("B15").value
                    # one = ws_out.range("C15").value
                    # ten = ws_out.range("D15").value

                    # print(f" Gp_pointone = {pointone}, Gp_one={one}, Gp_ten={ten}")

                    # Gp_pointone = ws_out.range("B15:B24").value
                    # Gp_one = ws_out.range("C15:C24").value
                    # Gp_ten = ws_out.range("D15:D24").value

                    # print(f" Gp_pointone = {Gp_pointone}, Gp_one={Gp_one}, Gp_ten={Gp_ten}")
                    cp.ok("\n ✅ Finished...")
                finally:
                    if wb_cel is not None:
                        wb_cel.close()
                    if app_cel is not None:
                        app_cel.quit()
        print(f"  Completed this batch in {time.time() - t_batch:.0f}s")

            
    

        


    except Exception as e:
        cp.err(f"\nError: {e}")
        import traceback
        traceback.print_exc()
        # don't leak the Excel session if we kept it open across steps
        try:
            if 'wb_xl' in locals() and wb_xl is not None:
                wb_xl.close()
            if 'app_xl' in locals() and app_xl is not None:
                app_xl.quit()
        except Exception:
            pass


