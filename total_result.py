"""
final_result.py — Scan for Excel outputs, read G' values, build grouped Prism.

Finds all *_bead*.xlsx files under the base directory, groups them by
condition (parent folder), reads G'(0.1/1/10 Hz) from the
"For GraphPad Graph&Stats" sheet, and creates a grouped .pzfx.
"""
# pip install openpyxl pandas prismWriter 

from pathlib import Path
import os
import sys
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
# from prismWriter.prism_writer import PrismFile 

# ══════════════════════════════════════════════════════════════
#  CONFIG — EDIT THESE PATHS
# ══════════════════════════════════════════════════════════════
# Every changeable path is listed here. Edit the values in this script.

# The experiment folder to scan and build results for.
BASE_DIR = "/Users/sharonnyoyoko/Downloads/redone anlysis"
# Example (Windows):  BASE_DIR = r"C:\Users\fntf0423\Downloads\Analysis\20251030 plasma new"

# Folder holding the Prism result template + the Excel / Prism bead templates.
AUTOMATION_REQUIREMENTS = "/Users/sharonnyoyoko/Downloads/redone anlysis/Automation Requirements"

# The Prism results template (the 9-table file carrying your graph).
# Absolute, or relative to AUTOMATION_REQUIREMENTS.
resulttemplate_path = os.path.join(AUTOMATION_REQUIREMENTS, "project_temp.pzfx")

# Output file naming (usually leave as-is).
BASE_NAME = os.path.basename(BASE_DIR.rstrip("/\\"))
OUTPUT_PZFX = os.path.join(BASE_DIR, f"{BASE_NAME}_results.pzfx")

# ---- Excel source sheet / ranges (usually leave as-is) ----
GP_SHEET_NAME = "For GraphPad Graph&Stats"
GP_POINTONE_RANGE = "B15:B24"
GP_ONE_RANGE = "C15:C24"
GP_TEN_RANGE = "D15:D24"
GP_SUMMARY_FILENAME = "Gp_summary.csv"



# ══════════════════════════════════════════════════════════════
#  PRISM NAMESPACE
# ══════════════════════════════════════════════════════════════

_NS = "http://graphpad.com/prism/Prism.htm"
_DT_NS = "urn:schemas-microsoft-com:datatypes"
ET.register_namespace("", _NS)
ET.register_namespace("dt", _DT_NS)

# tree = ET.parse(temp_pzfx)
# root = tree.getroot()


def ns(tag):
    return f"{{{_NS}}}{tag}".encode('utf-8').decode('utf-8')


def dt(attr):
    return f"{{{_DT_NS}}}{attr}"


# ══════════════════════════════════════════════════════════════
#  DISCOVERY
# ══════════════════════════════════════════════════════════════

def discover_summary_files(base_dir):
    base = Path(base_dir).expanduser().resolve()
    return sorted(base.rglob(GP_SUMMARY_FILENAME))


def discover_samples(base_dir):
    """Find all Excel output files and group by condition (parent folder).

    Looks for files matching *_bead*.xlsx under the base directory.
    The immediate parent folder name becomes the condition.
    The sample name is extracted from the filename stem.

    Returns: {condition: {sample_name: excel_path}}
    """
    base = Path(base_dir).expanduser().resolve()
    conditions = {}

    for xlsx in sorted(base.rglob("*_bead*.xlsx")):
        if xlsx.name.startswith("~$"):
            continue  # skip temp files
        condition = xlsx.parent.name
        # Extract sample name: "PD62_p4b2_bead2-11.xlsx" -> "p4b2"
        sample_name = xlsx.stem.split("_bead")[0]
        prefix = f"{condition}_"
        if sample_name.startswith(prefix):
            sample_name = sample_name[len(prefix):]
        conditions.setdefault(condition, {})[sample_name] = xlsx

    return conditions


# ══════════════════════════════════════════════════════════════
#  G' EXTRACTION
# ══════════════════════════════════════════════════════════════

def read_range(ws, cell_range):
    """Read a cell range and return flat list of values."""
    return [cell.value for row in ws[cell_range] for cell in row]


# grid block header in Gp_summary.csv -> key used by the final build
GRID_BLOCK_KEYS = {
    "G' 0.1Hz": "G'_0.1Hz",
    "G' 1Hz": "G'_1Hz",
    "G' 10Hz": "G'_10Hz",
    "G'' 0.1Hz": "G''_0.1Hz",
    "G'' 1Hz": "G''_1Hz",
    "G'' 10Hz": "G''_10Hz",
    "tanδ 0.1Hz": "tanδ_0.1Hz",
    "tanδ 1Hz": "tanδ_1Hz",
    "tanδ 10Hz": "tanδ_10Hz",
}


def load_gp_summary(summary_path):
    """Load one Gp_summary.csv and return its sample rows.

    Primary format (single file, grouped grid):
        two-level header (frequency block | 'condition sample'),
        rows = pull numbers, G' and G'' blocks side by side.
    Legacy wide Gp_values.csv (one row per sample) is also still accepted.

    Returns: list of (condition, sample_name, gp01, gp1, gp10,
                      gdp01, gdp1, gdp10, td01, td1, td10) tuples.
    """
    with open(summary_path, encoding="utf-8-sig") as f:
        first_line = f.readline()
    if "G' 0.1Hz" not in first_line:
        return load_gp_summary_wide(summary_path)  # legacy one-row-per-sample

    df = pd.read_csv(summary_path, encoding="utf-8-sig", header=[0, 1], index_col=0)
    by_sample = {}
    for col in df.columns:
        label, sample_col = str(col[0]), str(col[1])
        key = GRID_BLOCK_KEYS.get(label)
        if key is None or " " not in sample_col:
            continue
        condition, sample_name = sample_col.rsplit(" ", 1)
        vals = [v for v in df[col].dropna().tolist()]
        by_sample.setdefault((condition, sample_name), {})[key] = vals

    rows = []
    for (condition, sample_name), groups in by_sample.items():
        rows.append((condition, sample_name,
                     groups.get("G'_0.1Hz", []),
                     groups.get("G'_1Hz", []),
                     groups.get("G'_10Hz", []),
                     groups.get("G''_0.1Hz", []),
                     groups.get("G''_1Hz", []),
                     groups.get("G''_10Hz", []),
                     groups.get("tanδ_0.1Hz", []),
                     groups.get("tanδ_1Hz", []),
                     groups.get("tanδ_10Hz", [])))
    return rows


def load_gp_summary_wide(summary_path):
    """Legacy reader for the old wide Gp_values.csv (one row per sample)."""
    df = pd.read_csv(summary_path)
    rows = []

    def sort_gp_cols(cols):
        import re

        def sort_key(col):
            nums = re.findall(r"\d+", col)
            return [int(n) for n in nums] if nums else [col]

        return sorted(cols, key=sort_key)

    for _, row in df.iterrows():
        condition = str(row.get("condition", summary_path.parent.name))
        sample_name = row["sample_name"]

        # Extract G' values by matching both old and new summary column names.
        # Old format: gp01_1, gp01_2, ..., gp1_1, gp1_2, ..., gp10_1, gp10_2, ...
        # New format: gp01_123_pull123, gp1_123_pull123, gp10_123_pull123
        gp01_cols = sort_gp_cols([col for col in df.columns if col.startswith("gp01_") ])
        gp1_cols = sort_gp_cols([col for col in df.columns if col.startswith("gp1_") and not col.startswith("gp10_")])
        gp10_cols = sort_gp_cols([col for col in df.columns if col.startswith("gp10_") ])

        gp01 = [row.get(col) for col in gp01_cols]
        gp1 = [row.get(col) for col in gp1_cols]
        gp10 = [row.get(col) for col in gp10_cols]

        # G'' values too (legacy files may or may not have them)
        gdp01_cols = sort_gp_cols([col for col in df.columns if col.startswith("gdp01_")])
        gdp1_cols = sort_gp_cols([col for col in df.columns if col.startswith("gdp1_") and not col.startswith("gdp10_")])
        gdp10_cols = sort_gp_cols([col for col in df.columns if col.startswith("gdp10_")])

        gdp01 = [row.get(col) for col in gdp01_cols]
        gdp1 = [row.get(col) for col in gdp1_cols]
        gdp10 = [row.get(col) for col in gdp10_cols]

        # tanδ (legacy wide files usually won't have these)
        td01_cols = sort_gp_cols([col for col in df.columns if col.startswith("td01_")])
        td1_cols = sort_gp_cols([col for col in df.columns if col.startswith("td1_") and not col.startswith("td10_")])
        td10_cols = sort_gp_cols([col for col in df.columns if col.startswith("td10_")])

        td01 = [row.get(col) for col in td01_cols]
        td1 = [row.get(col) for col in td1_cols]
        td10 = [row.get(col) for col in td10_cols]

        rows.append((condition, sample_name, gp01, gp1, gp10, gdp01, gdp1, gdp10, td01, td1, td10))
    return rows


def tan_theta(gp, gdp):
    """tan theta = G'' / G' element-wise (None-safe, no division by zero)."""
    out = []
    for g, gd in zip(gp, gdp):
        if g is None or gd is None:
            out.append(None)
            continue
        try:
            g, gd = float(g), float(gd)
            out.append(gd / g if g != 0 else None)
        except Exception:
            out.append(None)
    return out


def load_gp_summaries(base_dir):
    """Load G', G'' and tanδ (G''/G') summary CSVs under base_dir.

    tanδ comes from the workbook-stored values in the grid (source of truth);
    when they are missing/empty, fall back to computing G''/G'."""
    summary_paths = discover_summary_files(base_dir)
    data = {"G'_0.1Hz": {}, "G'_1Hz": {}, "G'_10Hz": {},
            "G''_0.1Hz": {}, "G''_1Hz": {}, "G''_10Hz": {},
            "tanδ_0.1Hz": {}, "tanδ_1Hz": {}, "tanδ_10Hz": {}}
    for summary_path in summary_paths:
        for condition, sample_name, gp01, gp1, gp10, gdp01, gdp1, gdp10, td01, td1, td10 in load_gp_summary(summary_path):
            data["G'_0.1Hz"].setdefault(condition, {})[sample_name] = gp01
            data["G'_1Hz"].setdefault(condition, {})[sample_name] = gp1
            data["G'_10Hz"].setdefault(condition, {})[sample_name] = gp10
            data["G''_0.1Hz"].setdefault(condition, {})[sample_name] = gdp01
            data["G''_1Hz"].setdefault(condition, {})[sample_name] = gdp1
            data["G''_10Hz"].setdefault(condition, {})[sample_name] = gdp10
            data["tanδ_0.1Hz"].setdefault(condition, {})[sample_name] = td01 if td01 else tan_theta(gp01, gdp01)
            data["tanδ_1Hz"].setdefault(condition, {})[sample_name] = td1 if td1 else tan_theta(gp1, gdp1)
            data["tanδ_10Hz"].setdefault(condition, {})[sample_name] = td10 if td10 else tan_theta(gp10, gdp10)
    return data


def extract_gp_values(excel_path):
    """Read G'(0.1), G'(1), G'(10) from an Excel file.

    Returns: (gp01_list, gp1_list, gp10_list) — 10 or more values each
    """
    wb = load_workbook(excel_path, data_only=True)
    if GP_SHEET_NAME not in wb.sheetnames:
        wb.close()
        return None, None, None

    ws = wb[GP_SHEET_NAME]
    gp01 = read_range(ws, GP_POINTONE_RANGE)
    gp1 = read_range(ws, GP_ONE_RANGE)
    gp10 = read_range(ws, GP_TEN_RANGE)
    wb.close()
    return gp01, gp1, gp10

def sort_conditions(conditions):
    """Sort conditions based on a predefined order."""
    prefix_priority = {"AMI": 0, "PD": 1, "D": 2}
    
    def get_prefix(cond):
        # Extract prefix (letters only, before any numbers or underscores)
        prefix = ""
        for char in cond:
            if char.isalpha():
                prefix += char.upper()
            else:
                break
        return prefix
    
    return sorted(conditions, key=lambda x: (prefix_priority.get(get_prefix(x), 99), x))



def sort_nested(d):
    if not isinstance(d, dict):
        return d

    return {
        k: sort_nested(d[k])
        for k in sort_conditions(d.keys())
    }






# ══════════════════════════════════════════════════════════════
#  PRISM BUILDER
# ══════════════════════════════════════════════════════════════
#Didnt work well before. fixed in version two
def build_grouped_prism(data, times, output_path):
    """Create a .pzfx with grouped tables from a nested data dict.

    data = {
        "Gp_0.1Hz": {"AMI21": {"p1b2": [val,...], "p2b2": [...]}, ...},
        "Gp_1Hz":   {...},
        "Gp_10Hz":  {...},
    }
    """
    table_names = list(data.keys())
    if not table_names:
        raise ValueError("data dict is empty")

    first = data[table_names[0]]
    conditions = list(first.keys())
    sample_map = {c: list(first[c].keys()) for c in conditions}

    root = ET.Element(ns("GraphPadPrismFile"))
    root.set("PrismXMLVersion", "7.00")

    info = ET.SubElement(root, ns("Info"))
    info.set("Info table", "")
    info.set("Sequence", "1")
    info_table = ET.SubElement(info, ns("Table"))
    info_table.set("TableType", "Info")
    ET.SubElement(info_table, ns("Title")).text = "Info"

    sheet = ET.SubElement(root, ns("Sheet"))
    sheet.set("Sequence", "data")
    sheet.set("Name", "Data")

    for table_title, table_data in data.items():
        table_el = ET.SubElement(sheet, ns("Table"))
        table_el.set("TableType", "Grouped")
        table_el.set("XFormat", "none")
        ET.SubElement(table_el, ns("Title")).text = table_title

        rt = ET.SubElement(table_el, ns("RowTitlesColumn"))
        rt.set("Width", "60")
        rt.set("Decimals", "0")
        rsub = ET.SubElement(rt, ns("Subcolumn"))
        for t in times:
            d = ET.SubElement(rsub, ns("d"))
            d.text = str(t)

        for cond in conditions:
            samples = sample_map[cond]
            yc = ET.SubElement(table_el, ns("YColumn"))
            yc.set("Width", "81")
            yc.set("Subcolumns", str(len(samples)))
            yc.set("Decimals", "4")
            ET.SubElement(yc, ns("Title")).text = cond

            for sname in samples:
                sc = ET.SubElement(yc, ns("Subcolumn"))
                ET.SubElement(sc, ns("Title")).text = sname
                vals = table_data[cond][sname]
                if not isinstance(vals, (list, tuple)):
                    vals = [vals] * len(times)
                for v in vals:
                    d = ET.SubElement(sc, ns("d"))
                    d.text = str(v) if v is not None else ""

    root.set("xmlns:dt", _DT_NS)
    tree = ET.ElementTree(root)
    tree.write(output_path, xml_declaration=True, encoding="utf-8")

    # Post-process only line endings and XML declaration format
    with open(output_path, "r", encoding="utf-8") as f:
        content = f.read()

    content = content.replace(
        "<?xml version='1.0' encoding='utf-8'?>",
        '<?xml version="1.0" encoding="UTF-8"?>'
    )
    content = content.replace("\n", "\r\n")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"\nGrouped Prism saved -> {output_path}")
    print(f"  Tables: {table_names}")
    print(f"  Conditions: {conditions}")
    for c, s in sample_map.items():
        print(f"    {c}: {s}")

    return output_path

# ══════════════════════════════════════════════════════════════
# Instead of using prismWriter's make_group_table, we can directly write to the Prism XML tree of our template 
# so we can have graph output
# ══════════════════════════════════════════════════════════════

def build_grouped_prism_two(data, output_dir, stamp_notes=None, source_path=None):
    """Create a .pzfx with grouped tables from a nested data dict.

    data = {
        "Gp_0.1Hz": {"AMI21": {"p1b2": [val,...], "p2b2": [...]}, ...},
        "Gp_1Hz":   {...},
        "Gp_10Hz":  {...},
    }
    """
    table_names = list(data.keys())
    if not table_names:
        raise ValueError("data dict is empty")

    first = data[table_names[0]]
    conditions = list(first.keys())
    sample_map = {c: list(first[c].keys()) for c in conditions}

    print(f"""
    first: {first}
    conditions: {conditions}
    sample_map: {sample_map}""")

    tree = ET.parse(source_path if source_path else resulttemplate_path)
    root = tree.getroot()

    # use the TEMPLATE's actual XML namespace — old files use
    # "http://graphpad.com/prism/Prism.htm", Prism 10 uses
    # "http://graphpad.com/prism/Def/3.0/" (a hardcoded ns() would find nothing)
    _tpl_ns = root.tag.split('}')[0] + '}' if root.tag.startswith('{') else ''

    def ns(tag):
        return f"{_tpl_ns}{tag}"

    for table_title, table_data in data.items():
        # find the matching table in the template
        table_el = None
        for table in root.iter(ns("Table")):
            title_el = table.find(ns("Title"))
            if title_el is not None and (title_el.text or "").lower() == table_title.lower():
                table_el = table
                print(f"Found table {table_title} in template")
                break
        if table_el is None:
            print(f"  WARNING: Table '{table_title}' not found in template — skipping")
            continue

        # row titles (pull times) — replace any existing d elements; create
        # RowTitlesColumn if the template table lacks it (user-added tables
        # like Tanδ_* are bare shells with only a Title)
        rt = table_el.find(ns("RowTitlesColumn"))
        if rt is None:
            rt = ET.SubElement(table_el, ns("RowTitlesColumn"))
            rt.set("Width", "32")
        rsub = rt.find(ns("Subcolumn"))
        if rsub is None:
            rsub = ET.SubElement(rt, ns("Subcolumn"))
        max_length = 0
        for _cond, samples in table_data.items():
            for _sname, vals in samples.items():
                if len(vals) > max_length:
                    max_length = len(vals)
        for d in list(rsub.findall(ns("d"))):
            rsub.remove(d)
        for i in range(max_length):
            d = ET.SubElement(rsub, ns("d"))
            d.text = str(7 + 3 * i)

        # Collect existing YColumns so we can REUSE them in place — this keeps
        # the graph's data-set links alive across re-builds. In-place means
        # only the VALUES change; YColumn elements, order and IDs are kept.
        existing_yc = {}
        for old in table_el.findall(ns("YColumn")):
            t_el = old.find(ns("Title"))
            key = ((t_el.text or "").strip().lower()) if t_el is not None else ""
            existing_yc.setdefault(key, []).append(old)

        type_priority = {"AMI": 0, "PD": 1, "D": 2}

        def _type_of(cond):
            p = ""
            for ch in str(cond):
                if ch.isalpha():
                    p += ch.upper()
                else:
                    break
            return p or "OTHER"

        type_groups = {}
        for cond in table_data:
            t = _type_of(cond)
            for sname, vals in table_data[cond].items():
                type_groups.setdefault(t, []).append((cond, sname, vals))

        ordered_types = sorted(type_groups, key=lambda x: type_priority.get(x, 99))
        # Position i holds each type's i-th sample (empty where a type is
        # shorter). This matches the WORKING layout: SubColumnTitles has one
        # Subcolumn per position with ONE d per YColumn — that is how Prism
        # labels each group's subcolumns separately.
        total = max((len(type_groups[t]) for t in ordered_types), default=0)

        # Sync the template's Replicates (was 8) to the real subcolumn count,
        # otherwise Prism only displays the first 8 slots and PD/D look empty.
        table_el.set("Replicates", str(total))

        sct = table_el.find(ns("SubColumnTitles"))
        if sct is None:
            sct = ET.SubElement(table_el, ns("SubColumnTitles"))
            sct.set("OwnSet", "1")
        for sc in list(sct.findall(ns("Subcolumn"))):
            sct.remove(sc)
        for i in range(total):
            sc = ET.SubElement(sct, ns("Subcolumn"))
            for t in ordered_types:
                entries = type_groups[t]
                if i < len(entries):
                    c, s, _v = entries[i]
                    ET.SubElement(sc, ns("d")).text = f"{c} {s}"
                else:
                    ET.SubElement(sc, ns("d")).text = ""

        for t in ordered_types:
            entries = type_groups[t]
            pool = existing_yc.get(t.lower(), [])
            if pool:
                yc = pool.pop(0)                 # reuse element (graph link intact)
                title_el = yc.find(ns("Title"))
                if title_el is None:
                    title_el = ET.SubElement(yc, ns("Title"))
                title_el.text = t
                for sc in list(yc.findall(ns("Subcolumn"))):
                    yc.remove(sc)                # refill below
            else:
                yc = ET.SubElement(table_el, ns("YColumn"))
                yc.set("Width", "81")
                yc.set("Decimals", "4")
                ET.SubElement(yc, ns("Title")).text = t
            yc.set("Subcolumns", str(total))
            for i in range(total):
                sc = ET.SubElement(yc, ns("Subcolumn"))
                if i < len(entries):
                    _c, _s, vals = entries[i]
                    if not isinstance(vals, (list, tuple)):
                        vals = [vals] * max_length
                    for v in vals:
                        d = ET.SubElement(sc, ns("d"))
                        d.text = str(v) if v is not None else ""
                else:
                    for _r in range(max_length):
                        d = ET.SubElement(sc, ns("d"))
                        d.text = ""
        # empty any leftover YColumns (types no longer in the data) so the
        # graph keeps its slots but shows no stale values
        for leftovers in existing_yc.values():
            for yc in leftovers:
                for sc in list(yc.findall(ns("Subcolumn"))):
                    yc.remove(sc)
                for _i in range(total):
                    sc = ET.SubElement(yc, ns("Subcolumn"))
                    for _r in range(max_length):
                        ET.SubElement(sc, ns("d")).text = ""

        # root.set("xmlns:dt", _DT_NS)
        # tree = ET.ElementTree(root)
        # 

    output_path = os.path.join(output_dir, f"{BASE_NAME}_results.pzfx")

    if stamp_notes:
        for info in root.iter(ns("Info")):
            notes_el = info.find(ns("Notes"))
            if notes_el is not None:
                notes_el.text = ((notes_el.text or "").rstrip("\n") + "\n" + stamp_notes)
                break

    tree.write(output_path, xml_declaration=True, encoding="utf-8")

    # Post-process only line endings and XML declaration format
    with open(output_path, "r", encoding="utf-8") as f:
        content = f.read()

    content = content.replace(
        "<?xml version='1.0' encoding='utf-8'?>",
        '<?xml version="1.0" encoding="UTF-8"?>'
    )
    content = content.replace("\n", "\r\n")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"\nGrouped Prism saved -> {output_path}")
    print(f"  Tables: {table_names}")
    print(f"  Conditions: {conditions}")
    for c, s in sample_map.items():
        print(f"    {c}: {s}")

    return output_path


def flatten_data(data, times):
    """Convert nested data dict to a flat DataFrame with columns: Condiiton, Sample, Time, Value.
    Create a .pzfx with grouped tables from a nested data dict.

    data = {
        "G'_0.1Hz": {"AMI21": {"p1b2": [val,...], "p2b2": [...]}, ...},
        "G'_1Hz":   {...},
        "G'_10Hz":  {...},
    }
    """
    # table_names = list(data.keys())
    # if not table_names:
    #     raise ValueError("data dict is empty")

    rows = []
    for condition, subgroups in data.items():
        for subgroup, values in subgroups.items():
            padded_values = list(values) + [None] * max(0, len(times) - len(values))
            if len(padded_values) > len(times):
                padded_values = padded_values[:len(times)]

            for t, value in zip(times, padded_values):
                rows.append({
                    "condition": condition,
                    "subgroup": subgroup,
                    "time": t,
                    "value": value,
                    "decimals": 2
                })
    return pd.DataFrame(rows)

def sort_conditions(conditions):
    """Sort conditions based on a predefined order."""
    prefix_priority = {"AMI": 0, "PD": 1, "D": 2}
    
    def get_prefix(cond):
        # Extract prefix (letters only, before any numbers or underscores)
        prefix = ""
        for char in cond:
            if char.isalpha():
                prefix += char.upper()
            else:
                break
        return prefix
    
    return sorted(conditions, key=lambda x: (prefix_priority.get(get_prefix(x), 99), x))






def _data_hash(data):
    """Stable SHA-256 over the whole dataset (all table/condition/sample
    values) so you can tell when a build's data actually changed."""
    import hashlib
    parts = []
    for tbl in sorted(data):
        for cond in sorted(data[tbl]):
            for s in sorted(data[tbl][cond]):
                vals = []
                for v in data[tbl][cond][s]:
                    if v is None:
                        vals.append("")
                    elif isinstance(v, float) and v != v:
                        vals.append("nan")
                    else:
                        vals.append(repr(v))
                parts.append(f"{tbl}|{cond}|{s}|" + ",".join(vals))
    return hashlib.sha256("\n".join(parts).encode("utf-8")).hexdigest()


def write_build_manifest(output_path, stamp):
    """Write a sidecar <output>.build.json so you can verify which data a
    given pzfx contains without opening it."""
    import json
    manifest_path = output_path + ".build.json"
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(stamp, f, indent=2, ensure_ascii=False)
    return manifest_path


# order of the series — matches the Prism tables (G', G'', tanδ at each freq)
SERIES_ORDER = ["G'_0.1Hz", "G'_1Hz", "G'_10Hz",
                "G''_0.1Hz", "G''_1Hz", "G''_10Hz",
                "tanδ_0.1Hz", "tanδ_1Hz", "tanδ_10Hz"]


def write_final_excel(data, output_path):
    """Write one Excel workbook with all the data that goes into Prism.

    One sheet per series (G'/G''/tanδ at 0.1/1/10 Hz); each sheet is a grid
    with rows = pull times (7, 10, 13 ...) and columns = "condition sample".
    """
    import openpyxl
    from openpyxl.styles import Font
    wb = openpyxl.Workbook()
    first = True
    for series in SERIES_ORDER:
        series_data = data.get(series, {})
        if first:
            ws = wb.active
            ws.title = series[:31]
            first = False
        else:
            ws = wb.create_sheet(series[:31])
        pairs = [(cond, s) for cond in sorted(series_data)
                 for s in sorted(series_data[cond])]
        max_len = max((len(series_data[c][s]) for c in series_data
                       for s in series_data[c]), default=0)
        ws.cell(1, 1, "Pull")
        ws.cell(1, 1).font = Font(bold=True)
        for j, (cond, s) in enumerate(pairs, start=2):
            c = ws.cell(1, j, f"{cond} {s}")
            c.font = Font(bold=True)
        for i in range(max_len):
            ws.cell(i + 2, 1, 7 + 3 * i)
            for j, (cond, s) in enumerate(pairs, start=2):
                vals = series_data.get(cond, {}).get(s, [])
                if i < len(vals) and vals[i] is not None:
                    ws.cell(i + 2, j, vals[i])
    wb.save(output_path)
    return output_path


# ══════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":
    summary_files = discover_summary_files(BASE_DIR)
    if summary_files:
        print(f"Found {len(summary_files)} G' summary files. Reading summaries instead of opening Excel.")
        data = load_gp_summaries(BASE_DIR)
        conditions = {cond: list(samples.keys()) for cond, samples in data["G'_0.1Hz"].items()}
        total = sum(len(samples) for samples in conditions.values())
        print(f"Found {total} samples across {len(conditions)} conditions from summaries:\n")
        for cond, samples in sorted(conditions.items()):
            print(f"  {cond}/")
            for name in sorted(samples):
                print(f"    {name}")
    else:
        print("Scanning for Excel files...")
        conditions = discover_samples(BASE_DIR)

        if not conditions:
            print("No Excel files found. Check BASE_DIR.")
            sys.exit(1)

        total = sum(len(s) for s in conditions.values())
        print(f"Found {total} files across {len(conditions)} conditions:\n")
        for cond, samples in sorted(conditions.items()):
            print(f"  {cond}/")
            for name, path in sorted(samples.items()):
                print(f"    {name}  ->  {path.name}")

        # Read G' values
        data = {"G'_0.1Hz": {}, "G'_1Hz": {}, "G'_10Hz": {}}

        print(f"\nReading G' values...")
        for condition, samples in sorted(conditions.items()):
            for sample_name, excel_path in sorted(samples.items()):
                try:
                    gp01, gp1, gp10 = extract_gp_values(excel_path)
                    if gp01 is None:
                        print(f"  {condition}/{sample_name}: SKIP — '{GP_SHEET_NAME}' not found")
                        continue

                    data["G'_0.1Hz"].setdefault(condition, {})[sample_name] = gp01
                    data["G'_1Hz"].setdefault(condition, {})[sample_name] = gp1
                    data["G'_10Hz"].setdefault(condition, {})[sample_name] = gp10

                    print(f"  {condition}/{sample_name}: done")

                except Exception as e:
                    print(f"  {condition}/{sample_name}: ERROR — {e}")

        # Clean empties
        for tbl in list(data.keys()):
            data[tbl] = {k: v for k, v in data[tbl].items() if v}
            if not data[tbl]:
                del data[tbl]

        if not data:
            print("\nNo data collected.")
            sys.exit(1)

    sorted_data = sort_nested(data)

    # ── build stamp / change-detection ───────────────────────────────
    import datetime, json, shutil
    output_path = os.path.join(BASE_DIR, f"{BASE_NAME}_results.pzfx")
    stamp = {
        "output": output_path,
        "built_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "source": str(BASE_DIR),
        "tables": sorted(data.keys()),
        "samples": sum(len(s) for s in data["G'_0.1Hz"].values()),
        "conditions": {cond: sorted(samples) for cond, samples in data["G'_0.1Hz"].items()},
        "data_hash": _data_hash(data),
    }
    stamp_notes = (f"Auto-build {stamp['built_at']} from {stamp['source']} — "
                   f"{stamp['samples']} samples, {len(stamp['conditions'])} conditions, "
                   f"{len(stamp['tables'])} tables, hash {stamp['data_hash'][:12]}")

    # keep a recoverable copy of the previous (graph-connected) output
    if os.path.exists(output_path):
        try:
            shutil.copyfile(output_path, output_path.replace(".pzfx", ".previous.pzfx"))
        except Exception as exc:
            print(f"  ! could not back up previous output: {exc}")

    prev = {}
    try:
        with open(output_path + ".build.json", encoding="utf-8") as f:
            prev = json.load(f)
    except Exception:
        pass

    # Option 2: update the previous output IN PLACE (preserves the graph's
    # data-set links). The source is the previous output when it exists (so
    # the connected graph is carried forward); otherwise the pristine template
    # (first build). To force a fresh build from the template, delete the
    # output pzfx file before running.
    if os.path.exists(output_path):
        build_grouped_prism_two(sorted_data, BASE_DIR, stamp_notes=stamp_notes, source_path=output_path)
        in_place = True
    else:
        build_grouped_prism_two(sorted_data, BASE_DIR, stamp_notes=stamp_notes)
        in_place = False
    write_build_manifest(output_path, stamp)

    # final Excel with all the data that goes into Prism (one sheet per series)
    excel_out = os.path.join(BASE_DIR, f"{BASE_NAME}_final_data.xlsx")
    try:
        write_final_excel(data, excel_out)
    except Exception as exc:
        print(f"  ! could not write final Excel: {exc}")
        excel_out = None

    changed = prev.get("data_hash") != stamp["data_hash"]
    print("\n" + "=" * 62)
    print("BUILD COMPLETE")
    print(f"  output:    {output_path}")
    if excel_out:
        print(f"  excel:     {excel_out}")
    print(f"  built at:  {stamp['built_at']}")
    print(f"  source:    {stamp['source']}")
    print(f"  mode:      {'IN-PLACE — previous output updated, graph preserved' if in_place else 'FRESH — built from pristine template'}")
    print(f"  samples:   {stamp['samples']}  conditions: {len(stamp['conditions'])}  tables: {len(stamp['tables'])}")
    print(f"  data hash: {stamp['data_hash'][:16]}")
    if prev:
        print(f"  previous:  {prev.get('built_at')}  (hash {prev.get('data_hash', '')[:16]})")
        print(f"  status:    DATA {'CHANGED' if changed else 'UNCHANGED — same data as last build'}")
    else:
        print("  status:    first build (no previous manifest)")
    print("=" * 62)
    # Infer time points from the longest sample list.
    # all_lengths = {
    #     len(values)
    #     for experiment_data in data.values()
    #     for values in experiment_data.values()
    # }
    # sample_length = max(all_lengths)
    # if len(all_lengths) != 1:
    #     print(
    #         f"Warning: sample lengths differ across data sets {sorted(all_lengths)}. "
    #         f"Using max length {sample_length} and padding shorter series with None."
    #     )
    # # times = [7 + 3 * i for i in range(sample_length)]
    # times = [7, 10, 13, 16, 19, 22, 25, 28, 31, 34, 37, 40, 43, 46, 49, 52, 55, 58, 61, 64]

    # # Build
    # # build_grouped_prism(data, times, OUTPUT_PZFX)
    # pf = PrismFile()
    # # pf.load(resulttemplate_path)
    # # tables = pf.get_table_names()
    # # print(f"Loaded Prism template with tables: {tables}")


    

    # for experiment_name, experiment_data in data.items():
    #     df = flatten_data(experiment_data, times)

    #     ordered_conditions = sort_conditions(df['condition'].unique())
    #     condition_order = {cond: i for i, cond in enumerate(ordered_conditions)}
    #     df['_cond_order'] = df['condition'].map(condition_order)
    #     df = df.sort_values(['_cond_order', 'subgroup', 'time'])
    #     df = df.drop(columns=['_cond_order'])

    #     pf.make_group_table(
    #         group_name = experiment_name,
    #         group_values = df,
    #         groupby = "condition",
    #         subgroupby = "subgroup",
    #         rowgroupby = "time",
    #         cols = ["value"],
    #     )
    
    # pf.save(OUTPUT_PZFX)

    print("✅ Created fixed_times_subgroups.pzfx with subcolumns for each treatment.")